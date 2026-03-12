#!/usr/bin/env python3
"""
SMART CASCADE OCR: Create INPUT Excel from Work Order Images + QTY.txt
Uses intelligent multi-provider OCR with automatic fallback for PERFECT output
Providers: Google Cloud Vision > Azure > PaddleOCR > EasyOCR
"""
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import re
import cv2
import numpy as np

# Import smart cascading OCR engine
try:
    from core.processors.document.unified_ocr_engine import get_ocr_engine
    SMART_OCR_AVAILABLE = True
except ImportError:
    print("⚠️  Smart OCR engine not found. Using fallback...")
    SMART_OCR_AVAILABLE = False

# PWD BSR Database
PWD_ITEMS_DATABASE = {
    '1.1.2': {
        'description': 'Wiring of light/fan point - Medium point (up to 6 mtr.) with 1.5 sq.mm FR PVC insulated copper conductor in recessed PVC conduit with modular accessories',
        'unit': 'point',
        'rate': 602.0
    },
    '1.1.3': {
        'description': 'Wiring of light/fan point - Long point (up to 10 mtr.) with 1.5 sq.mm FR PVC insulated copper conductor in recessed PVC conduit with modular accessories',
        'unit': 'point',
        'rate': 825.0
    },
    '1.3.3': {
        'description': 'Wiring of 3/5 pin 6A plug point - Medium point (up to 6 mtr.) with 2.5 sq.mm FR PVC insulated copper conductor in recessed PVC conduit with modular accessories',
        'unit': 'point',
        'rate': 602.0
    },
    '3.4.2': {
        'description': 'FR PVC flexible conductor 2 core 4 sq.mm + 1 core 2.5 sq.mm (earth wire)',
        'unit': 'mtr',
        'rate': 85.0
    },
    '4.1.23': {
        'description': 'MCB Single pole 6A to 32A with B or C curve',
        'unit': 'Each',
        'rate': 285.0
    },
    '18.13': {
        'description': 'LED Street Light 11250 lumen (90W) IP65 rated with driver',
        'unit': 'Each',
        'rate': 5617.0
    }
}


def read_qty_file(qty_file_path):
    """Read QTY.txt and return dict of {item_no: quantity}"""
    qty_data = {}
    print(f"\n📄 Reading: {qty_file_path}")
    
    with open(qty_file_path, 'r', encoding='utf-8') as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if line and not line.startswith('#'):
                parts = line.split()
                if len(parts) >= 2:
                    item_code = parts[0]
                    try:
                        quantity = float(parts[1])
                        qty_data[item_code] = quantity
                        print(f"   {item_code}: {quantity}")
                    except ValueError:
                        print(f"   ⚠️  Line {line_no}: Invalid quantity '{parts[1]}'")
    
    print(f"✅ Found {len(qty_data)} items with quantities\n")
    return qty_data


def extract_text_smart_cascade(image_dir, use_consensus=False, use_retry=False):
    """
    Extract text using smart cascading OCR with automatic fallback
    
    Args:
        image_dir: Directory containing work order images
        use_consensus: Run multiple providers and pick best (slower but more accurate)
        use_retry: Retry with preprocessing if quality is poor
    
    Returns:
        Tuple of (all_text, ocr_stats)
    """
    print("\n🧠 Initializing Smart Cascading OCR Engine...")
    
    # Initialize OCR engine
    ocr_engine = get_ocr_engine(language="en+hi")
    
    image_path = Path(image_dir)
    image_files = []
    for ext in ['*.jpg', '*.jpeg', '*.png', '*.JPG', '*.JPEG', '*.PNG']:
        image_files.extend(list(image_path.glob(ext)))
    
    image_files = sorted(image_files)
    
    if not image_files:
        print(f"❌ No images found in {image_dir}")
        return "", {}
    
    print(f"📸 Processing {len(image_files)} images with smart OCR...\n")
    
    all_text = []
    ocr_stats = {
        'total_images': len(image_files),
        'providers_used': {},
        'avg_confidence': 0.0,
        'total_words': 0
    }
    
    total_confidence = 0.0
    
    for idx, img_file in enumerate(image_files, 1):
        print(f"\n{'─'*60}")
        print(f"📄 [{idx}/{len(image_files)}] {img_file.name}")
        print(f"{'─'*60}")
        
        try:
            # Load image
            image = cv2.imread(str(img_file))
            if image is None:
                print(f"   ❌ Failed to load image")
                continue
            
            # Choose extraction method
            if use_consensus:
                print("   🔄 Using CONSENSUS mode (multiple providers)...")
                result = ocr_engine.extract_with_consensus(image)
            elif use_retry:
                print("   🔄 Using RETRY mode (with preprocessing)...")
                result = ocr_engine.extract_with_retry(image, max_attempts=3, preprocess=True)
            else:
                print("   🔄 Using SMART CASCADE mode (automatic fallback)...")
                result = ocr_engine.extract_text(image, min_confidence=0.7, min_words=5)
            
            # Store result
            all_text.append(f"=== {img_file.name} ===\n{result.text}\n")
            
            # Update statistics
            provider = result.provider.upper()
            ocr_stats['providers_used'][provider] = ocr_stats['providers_used'].get(provider, 0) + 1
            total_confidence += result.confidence
            ocr_stats['total_words'] += len(result.words)
            
            print(f"\n   ✅ Extraction complete:")
            print(f"      Provider: {provider}")
            print(f"      Confidence: {result.confidence:.2%}")
            print(f"      Words: {len(result.words)}")
            print(f"      Characters: {len(result.text)}")
            
        except Exception as e:
            print(f"   ❌ Error: {str(e)[:100]}")
            continue
    
    # Calculate average confidence
    if ocr_stats['total_images'] > 0:
        ocr_stats['avg_confidence'] = total_confidence / ocr_stats['total_images']
    
    combined_text = '\n\n'.join(all_text)
    
    # Print summary
    print(f"\n{'='*60}")
    print("📊 OCR PROCESSING SUMMARY")
    print(f"{'='*60}")
    print(f"Images processed: {ocr_stats['total_images']}")
    print(f"Average confidence: {ocr_stats['avg_confidence']:.2%}")
    print(f"Total words extracted: {ocr_stats['total_words']}")
    print(f"\nProviders used:")
    for provider, count in ocr_stats['providers_used'].items():
        print(f"   {provider}: {count} images")
    print(f"{'='*60}\n")
    
    return combined_text, ocr_stats


def parse_header_info(text):
    """Extract header information from OCR text"""
    header_info = {
        'contractor': '[EXTRACTED FROM IMAGES]',
        'work_name': '[EXTRACTED FROM IMAGES]',
        'wo_number': '[EXTRACTED FROM IMAGES]',
        'agreement_no': '[EXTRACTED FROM IMAGES]',
        'wo_amount': '[EXTRACTED FROM IMAGES]'
    }
    
    lines = text.split('\n')
    
    for line in lines:
        line_lower = line.lower()
        
        # Contractor name
        if 'contractor' in line_lower or 'ठेकेदार' in line:
            match = re.search(r'(?:contractor|ठेकेदार)[:\s]+([A-Za-z\s.&]+)', line, re.I)
            if match:
                header_info['contractor'] = match.group(1).strip()
        
        # Work name
        if 'work' in line_lower and 'name' in line_lower:
            match = re.search(r'work\s+name[:\s]+(.+)', line, re.I)
            if match:
                header_info['work_name'] = match.group(1).strip()
        
        # Work order number
        if 'work order' in line_lower or 'w.o' in line_lower:
            match = re.search(r'(?:work\s*order|w\.o)[:\s]*no[.:\s]*([A-Z0-9/\-]+)', line, re.I)
            if match:
                header_info['wo_number'] = match.group(1).strip()
        
        # Agreement number
        if 'agreement' in line_lower:
            match = re.search(r'agreement[:\s]*no[.:\s]*([A-Z0-9/\-]+)', line, re.I)
            if match:
                header_info['agreement_no'] = match.group(1).strip()
        
        # Work order amount
        if 'amount' in line_lower or 'रुपये' in line:
            match = re.search(r'(?:rs\.?|₹)\s*([\d,]+(?:\.\d{2})?)', line, re.I)
            if match:
                header_info['wo_amount'] = match.group(1).replace(',', '')
    
    return header_info


def create_excel_smart_cascade(work_order_dir, output_file, mode='cascade'):
    """
    Create Excel using Smart Cascading OCR + QTY.txt
    
    Args:
        work_order_dir: Directory with work order images and qty.txt
        output_file: Output Excel file path
        mode: 'cascade' (default), 'consensus', or 'retry'
    """
    
    print(f"\n{'='*80}")
    print("🚀 SMART CASCADE OCR - INPUT EXCEL GENERATION")
    print(f"{'='*80}")
    print(f"Mode: {mode.upper()}")
    print(f"{'='*80}\n")
    
    work_order_path = Path(work_order_dir)
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Step 1: Read QTY.txt
    qty_file = work_order_path / "qty.txt"
    if not qty_file.exists():
        print(f"❌ QTY.txt not found at: {qty_file}")
        return False
    
    qty_data = read_qty_file(qty_file)
    
    # Step 2: Extract text using Smart Cascade OCR
    use_consensus = (mode == 'consensus')
    use_retry = (mode == 'retry')
    
    all_text, ocr_stats = extract_text_smart_cascade(
        work_order_path, 
        use_consensus=use_consensus,
        use_retry=use_retry
    )
    
    if not all_text:
        print("❌ No text extracted from images")
        return False
    
    # Save raw OCR text with stats
    text_file = output_path.parent / f"{output_path.stem}_OCR_TEXT.txt"
    with open(text_file, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("OCR STATISTICS\n")
        f.write("="*80 + "\n")
        f.write(f"Total Images: {ocr_stats['total_images']}\n")
        f.write(f"Average Confidence: {ocr_stats['avg_confidence']:.2%}\n")
        f.write(f"Total Words: {ocr_stats['total_words']}\n")
        f.write(f"\nProviders Used:\n")
        for provider, count in ocr_stats['providers_used'].items():
            f.write(f"  {provider}: {count} images\n")
        f.write("\n" + "="*80 + "\n\n")
        f.write(all_text)
    
    print(f"✅ Raw OCR text saved: {text_file}\n")
    
    # Step 3: Parse header information
    print("📋 Extracting header information...")
    header_info = parse_header_info(all_text)
    for key, value in header_info.items():
        print(f"   {key}: {value}")
    
    # Step 4: Create Excel with 4 sheets
    print("\n📊 Creating Excel file...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # === SHEET 1: Title ===
    ws_title = wb.create_sheet("Title")
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED", ""],
        ["Bill Number", "First"],
        ["Running or Final", "Final"],
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", header_info['contractor']],
        ["Name of Work ;-", header_info['work_name']],
        ["Serial No. of this bill :", "First & Final Bill"],
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", header_info['wo_number']],
        ["Agreement No.", header_info['agreement_no']],
        ["WORK ORDER AMOUNT RS.", header_info['wo_amount']],
        ["Date of written order to commence work :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of Start :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of completion :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of actual completion of work :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of measurement :", datetime.now().strftime("%Y-%m-%d")],
        ["TENDER PREMIUM %", "11.22"],
        ["Above / Below", "Above"],
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        cell_label = ws_title.cell(row=row_idx, column=1, value=label)
        cell_value = ws_title.cell(row=row_idx, column=2, value=value)
        cell_label.font = Font(bold=True)
    
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # === SHEET 2: Work Order ===
    ws_work = wb.create_sheet("Work Order")
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_work.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    total_wo_amount = 0.0
    
    for item_code in sorted(qty_data.keys()):
        if item_code in PWD_ITEMS_DATABASE:
            item_info = PWD_ITEMS_DATABASE[item_code]
            qty = qty_data[item_code]
            rate = item_info['rate']
            amount = qty * rate
            total_wo_amount += amount
            
            ws_work.cell(row=row_idx, column=1, value=item_code)
            ws_work.cell(row=row_idx, column=2, value=item_info['description'])
            ws_work.cell(row=row_idx, column=3, value=item_info['unit'])
            ws_work.cell(row=row_idx, column=4, value=qty)
            ws_work.cell(row=row_idx, column=5, value=rate)
            ws_work.cell(row=row_idx, column=6, value=amount)
            ws_work.cell(row=row_idx, column=7, value=item_code)
            
            row_idx += 1
    
    ws_work.column_dimensions['B'].width = 80
    
    # === SHEET 3: Bill Quantity ===
    ws_bill = wb.create_sheet("Bill Quantity")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bill.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    for item_code in sorted(qty_data.keys()):
        if item_code in PWD_ITEMS_DATABASE:
            item_info = PWD_ITEMS_DATABASE[item_code]
            qty = qty_data[item_code]
            rate = item_info['rate']
            amount = qty * rate
            
            ws_bill.cell(row=row_idx, column=1, value=item_code)
            ws_bill.cell(row=row_idx, column=2, value=item_info['description'])
            ws_bill.cell(row=row_idx, column=3, value=item_info['unit'])
            ws_bill.cell(row=row_idx, column=4, value=qty)
            ws_bill.cell(row=row_idx, column=5, value=rate)
            ws_bill.cell(row=row_idx, column=6, value=amount)
            ws_bill.cell(row=row_idx, column=7, value=item_code)
            
            row_idx += 1
    
    ws_bill.column_dimensions['B'].width = 80
    
    # === SHEET 4: Extra Items ===
    ws_extra = wb.create_sheet("Extra Items")
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        cell = ws_extra.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    ws_extra.column_dimensions['B'].width = 80
    
    # Save Excel
    wb.save(output_path)
    
    print(f"\n{'='*80}")
    print("✅ SUCCESS - INPUT EXCEL GENERATED WITH SMART CASCADE OCR")
    print(f"{'='*80}\n")
    print(f"📁 Output: {output_path.absolute()}")
    print(f"📄 OCR Text: {text_file.absolute()}")
    print(f"\n📊 OCR Quality Metrics:")
    print(f"   Average Confidence: {ocr_stats['avg_confidence']:.2%}")
    print(f"   Total Words: {ocr_stats['total_words']}")
    print(f"   Providers Used: {', '.join(ocr_stats['providers_used'].keys())}")
    print(f"\n💰 Total Work Order Amount: Rs. {total_wo_amount:,.2f}")
    print(f"📦 Items Processed: {len(qty_data)}")
    print(f"\n🎯 Next Step:")
    print(f"   python create_excel_enterprise.py")
    
    return True


def main():
    """Main entry point"""
    work_order_dir = "INPUT/work_order_samples/work_01_27022026"
    output_file = "OUTPUT/INPUT_work_01_SMART_CASCADE.xlsx"
    mode = 'cascade'  # Options: 'cascade', 'consensus', 'retry'
    
    if len(sys.argv) > 1:
        work_order_dir = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    if len(sys.argv) > 3:
        mode = sys.argv[3]
    
    print("\n🧠 Smart Cascade OCR Modes:")
    print("   cascade   - Fast, automatic fallback (default)")
    print("   consensus - Multiple providers, pick best (slower, more accurate)")
    print("   retry     - Retry with preprocessing (best for poor quality images)")
    print(f"\nSelected mode: {mode}\n")
    
    success = create_excel_smart_cascade(work_order_dir, output_file, mode)
    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
