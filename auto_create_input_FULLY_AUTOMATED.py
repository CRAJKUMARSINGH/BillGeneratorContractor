#!/usr/bin/env python3
"""
FULLY AUTOMATED: Create INPUT Excel from Work Order Images + QTY.txt
Uses EasyOCR (no Tesseract required) - 100% Python-based
"""
import sys
from pathlib import Path
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import re

# Try to import EasyOCR
try:
    import easyocr
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("⚠️  EasyOCR not installed. Installing now...")
    print("   This is a one-time setup (may take 2-3 minutes)")
    import subprocess
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "easyocr"])
        import easyocr
        OCR_AVAILABLE = True
        print("✅ EasyOCR installed successfully!")
    except Exception as e:
        print(f"❌ Failed to install EasyOCR: {e}")
        print("\nManual installation:")
        print("  pip install easyocr")
        sys.exit(1)

# PWD BSR Database (for matching items)
PWD_ITEMS_DATABASE = {
    '1.1.2': {
        'description': 'Wiring of light/fan point - Medium point (up to 6 mtr.) with 1.5 sq.mm FR PVC insulated copper conductor in recessed PVC conduit with modular accessories',
        'unit': 'point',
        'rate': 602.0
    },
    '1.1.3': {
        'description': 'Wiring of light/fan point - Long point (up to 10 mtr.) with 1.5 sq.mm FR PVC insulated copper conductor in recessed PVC conduit with modular accessories',
        'unit': 'point',
        'rate': 825.0
    },
    '1.3.3': {
        'description': 'Wiring of 3/5 pin 6A plug point - Medium point (up to 6 mtr.) with 2.5 sq.mm FR PVC insulated copper conductor in recessed PVC conduit with modular accessories',
        'unit': 'point',
        'rate': 602.0
    },
    '3.4.2': {
        'description': 'FR PVC flexible conductor 2 core 4 sq.mm + 1 core 2.5 sq.mm (earth wire)',
        'unit': 'mtr',
        'rate': 85.0
    },
    '4.1.23': {
        'description': 'MCB Single pole 6A to 32A with B or C curve',
        'unit': 'Each',
        'rate': 285.0
    },
    '18.13': {
        'description': 'LED Street Light 11250 lumen (90W) IP65 rated with driver',
        'unit': 'Each',
        'rate': 5617.0
    }
}

def read_qty_file(qty_file_path):
    """Read QTY.txt and return dict of {item_no: quantity}"""
    qty_data = {}
    # Avoid emojis in console output for Windows compatibility
    print(f"\nReading: {qty_file_path}")
    
    with open(qty_file_path, 'r', encoding='utf-8') as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if line and not line.startswith('#'):
                parts = line.split()
                if len(parts) >= 2:
                    item_code = parts[0]
                    try:
                        quantity = float(parts[1])
                        qty_data[item_code] = quantity
                        print(f"   {item_code}: {quantity}")
                    except ValueError:
                        print(f"   ⚠️  Line {line_no}: Invalid quantity '{parts[1]}'")
    
    print(f"Found {len(qty_data)} items with quantities\n")
    return qty_data

def extract_text_from_images_easyocr(image_dir):
    """Extract text from all images using EasyOCR"""
    # Avoid emojis in console output for Windows compatibility
    print("\nInitializing EasyOCR (English + Hindi)...")
    print("   First run will download models (~100MB)")
    
    reader = easyocr.Reader(['en', 'hi'], gpu=False)
    print("OCR engine ready\n")
    
    image_path = Path(image_dir)
    image_files = []
    for ext in ['*.jpg', '*.jpeg', '*.png', '*.JPG', '*.JPEG', '*.PNG']:
        image_files.extend(list(image_path.glob(ext)))
    
    image_files = sorted(image_files)
    
    if not image_files:
        print(f"No images found in {image_dir}")
        return ""
    
    print(f"Processing {len(image_files)} images...\n")
    
    all_text = []
    for idx, img_file in enumerate(image_files, 1):
        print(f"   [{idx}/{len(image_files)}] {img_file.name}")
        try:
            result = reader.readtext(str(img_file), detail=0)
            text = '\n'.join(result)
            all_text.append(f"=== {img_file.name} ===\n{text}\n")
            print(f"       Extracted {len(text)} characters")
        except Exception as e:
            safe_msg = str(e).encode("ascii", "replace").decode("ascii")
            print(f"       Error: {safe_msg}")
    
    return '\n\n'.join(all_text)

def parse_header_info(text):
    """Extract header information from OCR text"""
    header_info = {
        'contractor': '[EXTRACTED FROM IMAGES]',
        'work_name': '[EXTRACTED FROM IMAGES]',
        'wo_number': '[EXTRACTED FROM IMAGES]',
        'agreement_no': '[EXTRACTED FROM IMAGES]',
        'wo_amount': '[EXTRACTED FROM IMAGES]'
    }
    
    lines = text.split('\n')
    
    for line in lines:
        line_lower = line.lower()
        
        # Contractor name
        if 'contractor' in line_lower or 'ठेकेदार' in line:
            match = re.search(r'(?:contractor|ठेकेदार)[:\s]+([A-Za-z\s.&]+)', line, re.I)
            if match:
                header_info['contractor'] = match.group(1).strip()
        
        # Work name
        if 'work' in line_lower and 'name' in line_lower:
            match = re.search(r'work\s+name[:\s]+(.+)', line, re.I)
            if match:
                header_info['work_name'] = match.group(1).strip()
        
        # Work order number
        if 'work order' in line_lower or 'w.o' in line_lower:
            match = re.search(r'(?:work\s*order|w\.o)[:\s]*no[.:\s]*([A-Z0-9/\-]+)', line, re.I)
            if match:
                header_info['wo_number'] = match.group(1).strip()
        
        # Agreement number
        if 'agreement' in line_lower:
            match = re.search(r'agreement[:\s]*no[.:\s]*([A-Z0-9/\-]+)', line, re.I)
            if match:
                header_info['agreement_no'] = match.group(1).strip()
        
        # Work order amount
        if 'amount' in line_lower or 'रुपये' in line:
            match = re.search(r'(?:rs\.?|₹)\s*([\d,]+(?:\.\d{2})?)', line, re.I)
            if match:
                header_info['wo_amount'] = match.group(1).replace(',', '')
    
    return header_info

def create_excel_fully_automated(work_order_dir, output_file):
    """Create Excel using EasyOCR + QTY.txt - FULLY AUTOMATED"""
    
    print(f"\n{'='*80}")
    # Avoid emojis in console output for Windows compatibility
    print("FULLY AUTOMATED INPUT EXCEL GENERATION")
    print(f"{'='*80}\n")
    
    work_order_path = Path(work_order_dir)
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Step 1: Read QTY.txt
    qty_file = work_order_path / "qty.txt"
    if not qty_file.exists():
        print(f"❌ QTY.txt not found at: {qty_file}")
        return False
    
    qty_data = read_qty_file(qty_file)
    
    # Step 2: Extract text from images using EasyOCR
    all_text = extract_text_from_images_easyocr(work_order_path)
    
    if not all_text:
        print("❌ No text extracted from images")
        return False
    
    # Save raw OCR text
    text_file = output_path.parent / f"{output_path.stem}_OCR_TEXT.txt"
    with open(text_file, 'w', encoding='utf-8') as f:
        f.write(all_text)
    print(f"\n✅ Raw OCR text saved: {text_file}\n")
    
    # Step 3: Parse header information
    print("📋 Extracting header information...")
    header_info = parse_header_info(all_text)
    for key, value in header_info.items():
        print(f"   {key}: {value}")
    
    # Step 4: Create Excel with 4 sheets
    print("\n📊 Creating Excel file...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # === SHEET 1: Title ===
    ws_title = wb.create_sheet("Title")
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED", ""],
        ["Bill Number", "First"],
        ["Running or Final", "Final"],
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", header_info['contractor']],
        ["Name of Work ;-", header_info['work_name']],
        ["Serial No. of this bill :", "First & Final Bill"],
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", header_info['wo_number']],
        ["Agreement No.", header_info['agreement_no']],
        ["WORK ORDER AMOUNT RS.", header_info['wo_amount']],
        ["Date of written order to commence work :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of Start :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of completion :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of actual completion of work :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of measurement :", datetime.now().strftime("%Y-%m-%d")],
        ["TENDER PREMIUM %", "11.22"],
        ["Above / Below", "Above"],
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        cell_label = ws_title.cell(row=row_idx, column=1, value=label)
        cell_value = ws_title.cell(row=row_idx, column=2, value=value)
        cell_label.font = Font(bold=True)
    
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # === SHEET 2: Work Order ===
    ws_work = wb.create_sheet("Work Order")
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_work.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    total_wo_amount = 0.0
    
    for item_code in sorted(qty_data.keys()):
        if item_code in PWD_ITEMS_DATABASE:
            item_info = PWD_ITEMS_DATABASE[item_code]
            qty = qty_data[item_code]
            rate = item_info['rate']
            amount = qty * rate
            total_wo_amount += amount
            
            ws_work.cell(row=row_idx, column=1, value=item_code)
            ws_work.cell(row=row_idx, column=2, value=item_info['description'])
            ws_work.cell(row=row_idx, column=3, value=item_info['unit'])
            ws_work.cell(row=row_idx, column=4, value=qty)
            ws_work.cell(row=row_idx, column=5, value=rate)
            ws_work.cell(row=row_idx, column=6, value=amount)
            ws_work.cell(row=row_idx, column=7, value=item_code)
            
            row_idx += 1
    
    ws_work.column_dimensions['B'].width = 80
    
    # === SHEET 3: Bill Quantity (Same as Work Order for First & Final Bill) ===
    ws_bill = wb.create_sheet("Bill Quantity")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bill.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    for item_code in sorted(qty_data.keys()):
        if item_code in PWD_ITEMS_DATABASE:
            item_info = PWD_ITEMS_DATABASE[item_code]
            qty = qty_data[item_code]
            rate = item_info['rate']
            amount = qty * rate
            
            ws_bill.cell(row=row_idx, column=1, value=item_code)
            ws_bill.cell(row=row_idx, column=2, value=item_info['description'])
            ws_bill.cell(row=row_idx, column=3, value=item_info['unit'])
            ws_bill.cell(row=row_idx, column=4, value=qty)
            ws_bill.cell(row=row_idx, column=5, value=rate)
            ws_bill.cell(row=row_idx, column=6, value=amount)
            ws_bill.cell(row=row_idx, column=7, value=item_code)
            
            row_idx += 1
    
    ws_bill.column_dimensions['B'].width = 80
    
    # === SHEET 4: Extra Items (Empty) ===
    ws_extra = wb.create_sheet("Extra Items")
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        cell = ws_extra.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    ws_extra.column_dimensions['B'].width = 80
    
    # Save Excel
    wb.save(output_path)
    
    print(f"\n{'='*80}")
    print("✅ SUCCESS - INPUT EXCEL GENERATED")
    print(f"{'='*80}\n")
    print(f"📁 Output: {output_path.absolute()}")
    print(f"📄 OCR Text: {text_file.absolute()}")
    print(f"\n💰 Total Work Order Amount: Rs. {total_wo_amount:,.2f}")
    print(f"📦 Items Processed: {len(qty_data)}")
    print(f"\n🎯 Next Step:")
    print(f"   python create_excel_enterprise.py")
    print(f"   OR")
    print(f"   python process_first_bill.py {output_path}")
    
    return True

def main():
    """Main entry point"""
    work_order_dir = "INPUT/work_order_samples/work_01_27022026"
    output_file = "OUTPUT/INPUT_work_01_FULLY_AUTO.xlsx"
    
    if len(sys.argv) > 1:
        work_order_dir = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    success = create_excel_fully_automated(work_order_dir, output_file)
    return 0 if success else 1

if __name__ == '__main__':
    sys.exit(main())
