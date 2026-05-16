#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch Test: All 6 templates x 9 test files -> HTML + PDF outputs
Verifies:
  - Files with and without extra items
  - Extra items populate deviation statement (E-01, E-02...) with qty/amount
  - Extra items populate note sheet (extra_item_amount > 0 -> Yes)
  - All 6 templates render successfully
  - PDFs generated for each
"""
import sys
import json
import logging
from pathlib import Path
from datetime import datetime

# ── Path setup ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
ENGINE_DIR = ROOT / "engine"
sys.path.insert(0, str(ENGINE_DIR))

from calculation.excel_processor_enterprise import EnterpriseExcelProcessor
from calculation.bill_processor import process_bill
from rendering.html_renderer_enterprise import EnterpriseHTMLRenderer, RenderConfig, DocumentType
from rendering.pdf_generator import PDFGenerator
from model.document import BillDocument

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
UPLOADS_DIR = ROOT / "backend" / "uploads"
OUTPUT_BASE  = ROOT / "engine_output" / f"batch_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

SHEET_WO    = "WORK ORDER"
SHEET_BQ    = "BILL QUANTITY"
SHEET_EXTRA = "EXTRA ITEMS"

DOCUMENT_TYPES = [
    DocumentType.FIRST_PAGE,
    DocumentType.DEVIATION_STATEMENT,
    DocumentType.EXTRA_ITEMS,
    DocumentType.NOTE_SHEET,
    DocumentType.CERTIFICATE_II,
    DocumentType.CERTIFICATE_III,
]

TEMPLATE_VERSIONS = ["v1", "v2"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def load_sheets(xlsx: Path) -> dict:
    from openpyxl import load_workbook
    import pandas as pd

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    actual = {s.upper(): s for s in wb.sheetnames}
    wb.close()

    sheets_to_load = [v for k, v in actual.items() if k in (SHEET_WO, SHEET_BQ, SHEET_EXTRA, "TITLE")]
    proc = EnterpriseExcelProcessor(sanitize_strings=True, validate_schemas=False)
    result = proc.process_file(xlsx, sheet_names=sheets_to_load)
    if not result.success:
        raise RuntimeError(f"Excel load failed: {result.errors}")

    canonical = {}
    for canon in (SHEET_WO, SHEET_BQ, SHEET_EXTRA):
        real = actual.get(canon)
        if real and real in result.data:
            canonical[canon] = result.data[real]
    # Title sheet
    title_real = actual.get("TITLE")
    if title_real and title_real in result.data:
        canonical["TITLE"] = result.data[title_real]
    return canonical


_HEADER_MAP = {
    "agreement no": "agreement_no",
    "agreement no.": "agreement_no",
    "name of work": "name_of_work",
    "name of contractor or supplier": "name_of_firm",
    "name of contractor": "name_of_firm",
    "contractor": "name_of_firm",
    "work order amount rs.": "work_order_amount",
    "amount of work order": "work_order_amount",
    "date of written order to commence work": "date_commencement",
    "st. date of completion": "date_completion",
    "date of actual completion of work": "actual_completion",
}

def extract_meta(header_rows):
    meta = {v: "" for v in set(_HEADER_MAP.values())}
    meta["work_order_amount"] = 0.0
    for row in header_rows:
        cells = [str(c).strip() for c in row if str(c).strip() and str(c).strip() != "nan"]
        for i, cell in enumerate(cells):
            norm = cell.lower().rstrip(":;- ")
            field = _HEADER_MAP.get(norm)
            if field and i + 1 < len(cells):
                val = cells[i + 1]
                if field == "work_order_amount":
                    try:
                        meta[field] = float(str(val).replace(",", "").strip())
                    except ValueError:
                        pass
                elif not meta[field]:
                    meta[field] = val
    return meta


def build_doc(sheets, premium_percent=0.0, premium_type="above", prev_bill=0.0, source_filename=""):
    import pandas as pd
    ws_wo    = sheets.get(SHEET_WO)
    ws_bq    = sheets.get(SHEET_BQ)
    ws_extra = sheets.get(SHEET_EXTRA, pd.DataFrame())
    ws_title = sheets.get("TITLE")

    if ws_wo is None or ws_bq is None:
        raise RuntimeError(f"Missing required sheets. Found: {list(sheets.keys())}")

    fp, lp, dev, extra, ns = process_bill(
        ws_wo, ws_bq, ws_extra,
        premium_percent=premium_percent,
        premium_type=premium_type,
        previous_bill_amount=prev_bill,
    )
    meta = extract_meta(fp.get("header", []))
    extra_item_amount = float(fp["totals"].get("extra_items_sum", 0) or 0)

    # Parse Title sheet for v2 templates
    title_data = {}
    if ws_title is not None and not ws_title.empty:
        import numpy as np
        for _, row in ws_title.iterrows():
            cells = [str(c).strip() for c in row if str(c).strip() and str(c).strip() not in ("nan", "NaN")]
            if len(cells) >= 2:
                title_data[cells[0]] = cells[1]
    if not title_data:
        title_data = {
            "Agreement No.": meta["agreement_no"],
            "Name of Work ;-": meta["name_of_work"],
            "Name of Contractor or supplier :": meta["name_of_firm"],
            "Date of written order to commence work :": meta["date_commencement"],
            "St. date of completion :": meta["date_completion"],
            "Date of actual completion of work :": meta["actual_completion"],
        }

    return BillDocument(
        header=fp.get("header", []),
        items=fp.get("items", []),
        totals=fp.get("totals", {}),
        deviation_items=dev.get("items", []),
        deviation_summary=dev.get("summary", {}),
        extra_items=extra.get("items", []),
        agreement_no=meta["agreement_no"],
        name_of_work=meta["name_of_work"],
        name_of_firm=meta["name_of_firm"],
        date_commencement=meta["date_commencement"],
        date_completion=meta["date_completion"],
        actual_completion=meta["actual_completion"],
        work_order_amount=meta["work_order_amount"],
        extra_item_amount=extra_item_amount,
        title_data=title_data,
        source_filename=source_filename,
    )


def render_all(doc: BillDocument, out_dir: Path, version: str):
    tpl_dir = ENGINE_DIR / "templates" / version
    config = RenderConfig(template_dir=tpl_dir, output_dir=out_dir,
                          enable_security_checks=True, pdf_ready=True)
    renderer = EnterpriseHTMLRenderer(config)
    tdata = doc.to_template_dict()
    html_paths = []
    errors = []

    NOTE_SHEET_OVERRIDE = "note_sheet_new.html" if version == "v2" else None

    for dt in DOCUMENT_TYPES:
        fname = f"{dt.value}.html"

        # v2 note_sheet uses note_sheet_new.html
        if NOTE_SHEET_OVERRIDE and dt == DocumentType.NOTE_SHEET:
            try:
                tpl = renderer.template_manager.get_template(NOTE_SHEET_OVERRIDE)
                html_content = tpl.render(data=tdata)
                html_content = renderer._optimize_for_pdf(html_content)
                output_path = out_dir / fname
                output_path.write_text(html_content, encoding="utf-8")
                html_paths.append(output_path)
            except Exception as e:
                errors.append(f"{dt.value}: {e}")
            continue

        res = renderer.render(dt, {"data": tdata}, fname)
        if res.success:
            html_paths.append(res.output_path)
        else:
            errors.append(f"{dt.value}: {res.errors}")
    return html_paths, errors


def gen_pdfs(html_paths, out_dir: Path):
    gen = PDFGenerator(orientation="portrait")
    pdf_paths, pdf_errors = [], []
    for hp in html_paths:
        pp = out_dir / (hp.stem + ".pdf")
        try:
            engine = gen.generate_with_fallback(hp.read_text(encoding="utf-8"), str(pp))
            pdf_paths.append((pp, engine))
        except Exception as e:
            pdf_errors.append(f"{hp.name}: {e}")
    return pdf_paths, pdf_errors


def check_extra_items(doc: BillDocument):
    """Return dict of checks for extra item propagation."""
    checks = {}

    # 1. Are there actual extra items?
    real_extras = [i for i in doc.extra_items if not i.get("is_divider")]
    checks["has_extra_items"] = len(real_extras) > 0
    checks["extra_items_count"] = len(real_extras)

    # 2. extra_item_amount populated
    checks["extra_item_amount"] = doc.extra_item_amount
    checks["extra_item_amount_nonzero"] = doc.extra_item_amount > 0

    # 3. Deviation statement has extra items section
    dev_extras = [i for i in doc.deviation_items
                  if not i.get("is_divider") and str(i.get("serial_no", "")).startswith("E-")]
    checks["deviation_extra_count"] = len(dev_extras)
    checks["deviation_has_extras"] = len(dev_extras) > 0

    # 4. Deviation extra items have qty_bill and excess_amt populated
    if dev_extras:
        checks["deviation_extras_have_qty"] = all(
            (i.get("qty_bill") or 0) > 0 for i in dev_extras
        )
        checks["deviation_extras_have_excess_amt"] = all(
            (i.get("excess_amt") or 0) > 0 for i in dev_extras
        )
    else:
        checks["deviation_extras_have_qty"] = None
        checks["deviation_extras_have_excess_amt"] = None

    # 5. First page items include extra items divider + E- items
    fp_extras = [i for i in doc.items
                 if not i.get("is_divider") and str(i.get("serial_no", "")).startswith("E-")]
    checks["first_page_extra_count"] = len(fp_extras)

    # 6. totals.extra_items_sum
    checks["totals_extra_items_sum"] = doc.totals.get("extra_items_sum", 0)

    return checks


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    xlsx_files = sorted(UPLOADS_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print("[FAIL] No .xlsx files found in backend/uploads/")
        sys.exit(1)

    print(f"\n{'='*70}")
    print(f"BATCH TEST: {len(xlsx_files)} files x {len(TEMPLATE_VERSIONS)} template versions")
    print(f"Output: {OUTPUT_BASE}")
    print(f"{'='*70}\n")

    summary = []
    total_pass = total_fail = 0

    for xlsx in xlsx_files:
        file_label = xlsx.name[:20] + "..." if len(xlsx.name) > 20 else xlsx.name
        print(f"\n{'-'*70}")
        print(f"FILE: {xlsx.name}")

        # Load once
        try:
            sheets = load_sheets(xlsx)
        except Exception as e:
            print(f"  [FAIL] LOAD FAILED: {e}")
            total_fail += 1
            summary.append({"file": xlsx.name, "status": "LOAD_FAILED", "error": str(e)})
            continue

        has_extra_sheet = SHEET_EXTRA in sheets
        print(f"  Sheets loaded: {list(sheets.keys())}")
        print(f"  Extra Items sheet: {'YES' if has_extra_sheet else 'NO'}")

        # Build document once (shared across template versions)
        try:
            doc = build_doc(sheets, premium_percent=22.22, premium_type="above",
                            source_filename=xlsx.name)
        except Exception as e:
            print(f"  [FAIL] PROCESS FAILED: {e}")
            total_fail += 1
            summary.append({"file": xlsx.name, "status": "PROCESS_FAILED", "error": str(e)})
            continue

        # Extra items checks
        checks = check_extra_items(doc)
        print(f"\n  EXTRA ITEMS ANALYSIS:")
        print(f"    Has extra items:          {checks['has_extra_items']} ({checks['extra_items_count']} items)")
        print(f"    extra_item_amount:         {checks['extra_item_amount']}")
        print(f"    totals.extra_items_sum:    {checks['totals_extra_items_sum']}")
        print(f"    First page E- items:       {checks['first_page_extra_count']}")
        print(f"    Deviation E- items:        {checks['deviation_extra_count']}")
        if checks["has_extra_items"]:
            print(f"    Deviation qty populated:   {checks['deviation_extras_have_qty']}")
            print(f"    Deviation excess_amt:      {checks['deviation_extras_have_excess_amt']}")

        # Validate extra item propagation
        if checks["has_extra_items"]:
            ok = (
                checks["extra_item_amount_nonzero"] and
                checks["deviation_has_extras"] and
                checks["deviation_extras_have_qty"] and
                checks["deviation_extras_have_excess_amt"] and
                checks["first_page_extra_count"] > 0
            )
            print(f"    [OK] Extra items propagated correctly" if ok else
                  f"    [WARN] Extra items propagation ISSUES detected")
        else:
            print(f"    [INFO] No extra items -- NIL path")

        # Render for each template version
        for version in TEMPLATE_VERSIONS:
            out_dir = OUTPUT_BASE / xlsx.stem / version
            out_dir.mkdir(parents=True, exist_ok=True)

            print(f"\n  TEMPLATE {version.upper()}:")

            # HTML
            try:
                html_paths, html_errors = render_all(doc, out_dir, version)
                if html_errors:
                    for e in html_errors:
                        print(f"    [WARN] HTML error: {e}")
                print(f"    HTML: {len(html_paths)}/{len(DOCUMENT_TYPES)} rendered -> {out_dir.name}/")
            except Exception as e:
                print(f"    [FAIL] HTML render FAILED: {e}")
                html_paths = []

            # PDF
            if html_paths:
                try:
                    pdf_paths, pdf_errors = gen_pdfs(html_paths, out_dir)
                    if pdf_errors:
                        for e in pdf_errors:
                            print(f"    [WARN] PDF error: {e}")
                    for pp, engine in pdf_paths:
                        size_kb = pp.stat().st_size // 1024
                        print(f"    PDF [{engine:12s}] {pp.name:35s} {size_kb:>4} KB")
                    print(f"    PDF: {len(pdf_paths)}/{len(html_paths)} generated")
                except Exception as e:
                    print(f"    [FAIL] PDF generation FAILED: {e}")
                    pdf_paths = []
            else:
                pdf_paths = []

            file_ok = len(html_paths) == len(DOCUMENT_TYPES) and len(pdf_paths) == len(html_paths)
            if file_ok:
                total_pass += 1
            else:
                total_fail += 1

            summary.append({
                "file": xlsx.name,
                "version": version,
                "html_ok": len(html_paths),
                "pdf_ok": len(pdf_paths),
                "has_extra": checks["has_extra_items"],
                "extra_count": checks["extra_items_count"],
                "extra_amount": checks["extra_item_amount"],
                "deviation_extras": checks["deviation_extra_count"],
                "status": "PASS" if file_ok else "PARTIAL",
            })

    # ── Final Summary ─────────────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print("BATCH TEST SUMMARY")
    print(f"{'='*70}")
    print(f"{'FILE':<45} {'VER':<4} {'HTML':>5} {'PDF':>5} {'EXTRA':>6} {'STATUS'}")
    print(f"{'-'*70}")
    for r in summary:
        if "version" not in r:
            print(f"{r['file']:<45} {'':4} {'':5} {'':5} {'':6} {r['status']}")
        else:
            extra_str = f"{r['extra_count']}x" if r["has_extra"] else "NIL"
            print(f"{r['file'][:44]:<45} {r['version']:<4} {r['html_ok']:>5} {r['pdf_ok']:>5} {extra_str:>6} {r['status']}")

    print(f"\nTotal runs: {total_pass + total_fail}  PASS: {total_pass}  FAIL/PARTIAL: {total_fail}")
    print(f"Output dir: {OUTPUT_BASE}")

    # Save JSON report
    report_path = OUTPUT_BASE / "test_report.json"
    OUTPUT_BASE.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w") as f:
        json.dump(summary, f, indent=2, default=str)
    print(f"Report:     {report_path}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
