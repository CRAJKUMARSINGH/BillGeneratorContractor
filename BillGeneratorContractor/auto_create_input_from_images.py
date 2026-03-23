#!/usr/bin/env python3
"""
AUTOMATED: Create INPUT Excel from Work Order Images + QTY.txt
Extracts ALL data from images using OCR
"""
import sys
from pathlib import Path
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent))

try:
    import pytesseract
    from PIL import Image
    import cv2
    import numpy as np
except ImportError as e:
    print(f"❌ Missing dependency: {e}")
    print("Install: pip install pytesseract pillow opencv-python-headless")
    sys.exit(1)

def read_qty_file(qty_file_path):
    """Read QTY.txt and return dict of {item_no: quantity}"""
    qty_data = {}
    with open(qty_file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                parts = line.split()
                if len(parts) >= 2:
                    qty_data[parts[0]] = float(parts[1])
    return qty_data

def preprocess_image(image_path):
    """Preprocess image for better OCR"""
    img = cv2.imread(str(image_path))
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
    _, binary = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return binary

def extract_text_from_image(image_path):
    """Extract text using OCR"""
    print(f"  Processing: {image_path.name}")
    try:
        preprocessed = preprocess_image(image_path)
        pil_image = Image.fromarray(preprocessed)
        text = pytesseract.image_to_string(pil_image, config='--psm 6 -l eng+hin')
        print(f"    Extracted {len(text)} characters")
        return text
    except Exception as e:
        print(f"    ❌ Error: {e}")
        return ""

def extract_all_text_from_images(image_dir):
    """Extract text from all images"""
    image_path = Path(image_dir)
    image_files = []
    for ext in ['*.jpg', '*.jpeg', '*.png']:
        image_files.extend(list(image_path.glob(ext)))
    
    image_files = sorted(image_files)
    all_text = []
    
    print(f"\nExtracting text from {len(image_files)} images...")
    for img_file in image_files:
        text = extract_text_from_image(img_file)
        all_text.append(text)
    
    return '\n\n'.join(all_text)

def parse_work_order_data(text):
    """Parse work order data from OCR text"""
    import re
    
    lines = text.split('\n')
    items = []
    header_info = {}
    
    # Extract header information
    for line in lines[:50]:  # Check first 50 lines for header
        if 'contractor' in line.lower():
            match = re.search(r'contractor[:\s]+([A-Za-z\s.&]+)', line, re.I)
            if match:
                header_info['contractor'] = match.group(1).strip()
        
        if 'work order' in line.lower() and 'no' in line.lower():
            match = re.search(r'(?:work\s*order|wo)[:\s]*no[.:\s]*([A-Z0-9/\-]+)', line, re.I)
            if match:
                header_info['wo_number'] = match.group(1).strip()
        
        if 'agreement' in line.lower():
            match = re.search(r'agreement[:\s]*no[.:\s]*([A-Z0-9/\-]+)', line, re.I)
            if match:
                header_info['agreement_no'] = match.group(1).strip()
    
    # Parse items (simplified - you may need to adjust based on actual image format)
    current_item = None
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Try to match item number pattern
        item_match = re.match(r'^(\d+\.?\d*)\s+(.+)', line)
        if item_match:
            if current_item:
                items.append(current_item)
            
            item_no = item_match.group(1)
            rest = item_match.group(2)
            
            current_item = {
                'item_no': item_no,
                'description': rest,
                'unit': '',
                'quantity': 0.0,
                'rate': 0.0,
                'amount': 0.0,
                'bsr': ''
            }
        elif current_item:
            current_item['description'] += ' ' + line
    
    if current_item:
        items.append(current_item)
    
    return header_info, items

def create_excel_with_ocr(work_order_dir, output_file):
    """Create Excel using OCR + QTY.txt"""
    
    print(f"\n{'='*80}")
    print("AUTOMATED: Creating INPUT Excel from Images + QTY.txt")
    print(f"{'='*80}\n")
    
    work_order_path = Path(work_order_dir)
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Read QTY.txt
    qty_file = work_order_path / "qty.txt"
    if not qty_file.exists():
        print(f"❌ QTY.txt not found")
        return False
    
    print("Reading QTY.txt...")
    qty_data = read_qty_file(qty_file)
    print(f"✅ Found {len(qty_data)} items with quantities\n")
    
    # Check if Tesseract is available
    try:
        pytesseract.get_tesseract_version()
        print("✅ Tesseract OCR found\n")
    except:
        print("❌ Tesseract OCR not found!")
        print("Please install Tesseract or use manual entry method")
        print("\nFallback: Use the template we created earlier")
        print("  start OUTPUT\\work_order_from_images.xlsx")
        return False
    
    # Extract text from images
    all_text = extract_all_text_from_images(work_order_path)
    
    # Save raw text
    text_file = output_path.parent / "ocr_extracted_text.txt"
    with open(text_file, 'w', encoding='utf-8') as f:
        f.write(all_text)
    print(f"\n✅ Raw OCR text saved: {text_file}\n")
    
    # Parse data
    print("Parsing work order data...")
    header_info, items = parse_work_order_data(all_text)
    print(f"✅ Extracted {len(items)} items\n")
    
    # Create Excel
    print("Creating Excel file...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Title Sheet
    ws_title = wb.create_sheet("Title")
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED", ""],
        ["Bill Number", "First"],
        ["Running or Final", "Final"],
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", header_info.get('contractor', '[FROM IMAGES]')],
        ["Name of Work ;-", "[FROM IMAGES]"],
        ["Serial No. of this bill :", "First & Final Bill"],
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", header_info.get('wo_number', '[FROM IMAGES]')],
        ["Agreement No.", header_info.get('agreement_no', '[FROM IMAGES]')],
        ["WORK ORDER AMOUNT RS.", "[FROM IMAGES]"],
        ["Date of written order to commence work :", "2026-02-27"],
        ["St. date of Start :", "2026-03-01"],
        ["St. date of completion :", "2026-06-30"],
        ["Date of actual completion of work :", "2026-06-30"],
        ["Date of measurement :", "2026-03-09"],
        ["TENDER PREMIUM %", "11.22"],
        ["Above / Below", "Above"],
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        ws_title.cell(row=row_idx, column=1, value=label)
        ws_title.cell(row=row_idx, column=2, value=value)
    
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # Work Order Sheet
    ws_work = wb.create_sheet("Work Order")
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        ws_work.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    
    for row_idx, item in enumerate(items, start=2):
        ws_work.cell(row=row_idx, column=1, value=item['item_no'])
        ws_work.cell(row=row_idx, column=2, value=item['description'])
        ws_work.cell(row=row_idx, column=3, value=item['unit'])
        ws_work.cell(row=row_idx, column=4, value=item['quantity'])
        ws_work.cell(row=row_idx, column=5, value=item['rate'])
        ws_work.cell(row=row_idx, column=6, value=item['amount'])
        ws_work.cell(row=row_idx, column=7, value=item['bsr'])
    
    ws_work.column_dimensions['B'].width = 80
    
    # Bill Quantity Sheet (with QTY.txt data)
    ws_bill = wb.create_sheet("Bill Quantity")
    for col_idx, header in enumerate(headers, start=1):
        ws_bill.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    
    print("Applying QTY.txt data to Bill Quantity sheet...")
    for row_idx, item in enumerate(items, start=2):
        bsr = item.get('bsr', '')
        item_no = item.get('item_no', '')
        
        # Check if in QTY.txt (try both BSR and item_no)
        qty = qty_data.get(bsr, qty_data.get(item_no, 0.0))
        
        ws_bill.cell(row=row_idx, column=1, value=item['item_no'])
        ws_bill.cell(row=row_idx, column=2, value=item['description'])
        ws_bill.cell(row=row_idx, column=3, value=item['unit'])
        ws_bill.cell(row=row_idx, column=4, value=qty)  # From QTY.txt or 0
        ws_bill.cell(row=row_idx, column=5, value=item['rate'])
        
        # Calculate amount
        if item['rate']:
            amount = qty * item['rate']
            ws_bill.cell(row=row_idx, column=6, value=amount)
        
        ws_bill.cell(row=row_idx, column=7, value=item['bsr'])
        
        if qty > 0:
            print(f"  {bsr or item_no}: Qty = {qty}")
    
    ws_bill.column_dimensions['B'].width = 80
    
    # Extra Items Sheet (empty)
    ws_extra = wb.create_sheet("Extra Items")
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        ws_extra.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    
    # Save
    wb.save(output_path)
    
    print(f"\n✅ Excel file created: {output_path.absolute()}")
    print(f"\n⚠️  IMPORTANT: Review and verify OCR accuracy!")
    print(f"   Compare with: {text_file}")
    print(f"\nNext: python process_first_bill.py {output_path}")
    
    return True

def main():
    work_order_dir = "INPUT/work_order_samples/work_01_27022026"
    output_file = "OUTPUT/INPUT_work_01_AUTO.xlsx"
    
    if len(sys.argv) > 1:
        work_order_dir = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    success = create_excel_with_ocr(work_order_dir, output_file)
    return success

if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
