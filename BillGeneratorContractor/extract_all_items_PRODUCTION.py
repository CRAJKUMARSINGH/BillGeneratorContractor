#!/usr/bin/env python3
"""
PRODUCTION VERSION - Extract ALL items from images with improvements:
- Retry logic with exponential backoff
- Data validation
- File lock handling
- Progress tracking
"""
import sys
import os
import re
import time
from pathlib import Path
from datetime import datetime
sys.path.insert(0, str(Path(__file__).parent))

from modules.gemini_vision_parser_v2 import GeminiVisionParserV2
import openpyxl
from openpyxl.styles import Font, Alignment

def validate_item(item):
    """Validate extracted item data"""
    # Validate BSR code format
    if not re.match(r'^\d+\.\d+', item['code']):
        print(f"  WARNING: Invalid BSR code format: {item['code']}")
        return False
    
    # Validate rate range
    if item['rate'] <= 0 or item['rate'] > 100000:
        print(f"  WARNING: Invalid rate: {item['rate']} for {item['code']}")
        return False
    
    # Validate quantity range
    if item.get('quantity', 0) < 0 or item.get('quantity', 0) > 10000:
        print(f"  WARNING: Invalid quantity: {item.get('quantity')} for {item['code']}")
        return False
    
    return True

def is_file_locked(filepath):
    """Check if file is locked by another process"""
    try:
        with open(filepath, 'a'):
            return False
    except (IOError, PermissionError):
        return True

def save_excel_safely(wb, output_path):
    """Save Excel file with lock detection"""
    if output_path.exists() and is_file_locked(output_path):
        # Generate timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = output_path.stem
        new_path = output_path.parent / f"{base}_{timestamp}.xlsx"
        wb.save(new_path)
        print(f"\nOriginal file locked. Saved as: {new_path.name}")
        return new_path
    else:
        wb.save(output_path)
        return output_path

def main():
    # Set API key
    os.environ['GEMINI_API_KEY'] = 'AIzaSyBMZYPgjcqXY-tpe6UhtBtrWhzfbU0-WVU'
    
    work_dir = Path("INPUT_WORK_ORDER_IMAGES_TEXT")
    output_file = Path("OUTPUT/INPUT_FINAL_WITH_QUANTITIES.xlsx")
    
    print("="*80)
    print("PRODUCTION: EXTRACTING ALL ITEMS FROM ALL IMAGES")
    print("="*80)
    
    # Initialize parser
    parser = GeminiVisionParserV2()
    
    if not parser.available:
        print("\nGemini not available")
        return 1
    
    # Extract from ALL images with retry
    print("\nProcessing all images (with retry logic)...")
    all_items = parser.extract_all(work_dir)
    
    if not all_items:
        print("\nNo items extracted")
        return 1
    
    print(f"\nExtracted {len(all_items)} items from images")
    
    # VALIDATE extracted items
    print("\nValidating extracted items...")
    valid_items = []
    invalid_count = 0
    
    for item in all_items:
        if validate_item(item):
            valid_items.append(item)
        else:
            invalid_count += 1
    
    if invalid_count > 0:
        print(f"WARNING: {invalid_count} items failed validation and were removed")
    
    print(f"Valid items: {len(valid_items)}")
    
    # SORT by BSR code (not image order)
    def sort_key(item):
        code = item['code']
        parts = code.split('.')
        try:
            return tuple(int(p) for p in parts)
        except:
            return (999, 999, 999)
    
    valid_items.sort(key=sort_key)
    print("Items sorted by BSR code")
    
    # Read qty.txt
    qty_data = {}
    qty_file = work_dir / "qty.txt"
    if qty_file.exists():
        with open(qty_file, 'r') as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 2:
                    qty_data[parts[0]] = float(parts[1])
    
    # Create mapping for partial BSR code matches
    def find_qty_for_code(code):
        # Exact match first
        if code in qty_data:
            return qty_data[code]
        # Check if any qty code is a prefix of this code
        for qty_code, qty in qty_data.items():
            if code.startswith(qty_code + '.'):
                print(f"  Matched {qty_code} -> {code}")
                return qty
        return 0
    
    # Create Excel
    print("\nCreating Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Title sheet
    ws_title = wb.create_sheet("Title")
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED", ""],
        ["Bill Number", "First"],
        ["Running or Final", "Final"],
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", "[FROM IMAGES]"],
        ["Name of Work ;-", "[FROM IMAGES]"],
        ["Serial No. of this bill :", "First & Final Bill"],
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", "[FROM IMAGES]"],
        ["Agreement No.", "[FROM IMAGES]"],
        ["WORK ORDER AMOUNT RS.", "[FROM IMAGES]"],
        ["Date of written order to commence work :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of Start :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of completion :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of actual completion of work :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of measurement :", datetime.now().strftime("%Y-%m-%d")],
        ["TENDER PREMIUM %", "11.22"],
        ["Above / Below", "Above"],
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        ws_title.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_title.cell(row=row_idx, column=2, value=value)
    
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # Work Order sheet - WITH quantities and amounts from images
    ws_work = wb.create_sheet("Work Order")
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_work.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    total = 0.0
    for item in valid_items:
        # Work Order has quantities FROM IMAGES (if available)
        qty_from_image = item.get('quantity', 0)
        rate = item['rate']
        amount = qty_from_image * rate if qty_from_image > 0 else 0
        
        ws_work.cell(row=row_idx, column=1, value=item['code'])
        ws_work.cell(row=row_idx, column=2, value=item['description'])
        ws_work.cell(row=row_idx, column=3, value=item['unit'])
        ws_work.cell(row=row_idx, column=4, value=qty_from_image if qty_from_image > 0 else '')
        ws_work.cell(row=row_idx, column=5, value=rate)
        ws_work.cell(row=row_idx, column=6, value=amount if amount > 0 else '')
        ws_work.cell(row=row_idx, column=7, value=item['code'])
        
        if amount > 0:
            total += amount
        row_idx += 1
    
    ws_work.column_dimensions['B'].width = 80
    
    # Bill Quantity sheet
    ws_bill = wb.create_sheet("Bill Quantity")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bill.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    bill_total = 0.0
    for item in valid_items:
        code = item['code']
        qty = find_qty_for_code(code)
        
        if qty > 0:
            amount = qty * item['rate']
            ws_bill.cell(row=row_idx, column=1, value=code)
            ws_bill.cell(row=row_idx, column=2, value=item['description'])
            ws_bill.cell(row=row_idx, column=3, value=item['unit'])
            ws_bill.cell(row=row_idx, column=4, value=qty)
            ws_bill.cell(row=row_idx, column=5, value=item['rate'])
            ws_bill.cell(row=row_idx, column=6, value=amount)
            ws_bill.cell(row=row_idx, column=7, value=code)
            bill_total += amount
            row_idx += 1
    
    ws_bill.column_dimensions['B'].width = 80
    
    # Extra Items sheet
    ws_extra = wb.create_sheet("Extra Items")
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        cell = ws_extra.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    ws_extra.column_dimensions['B'].width = 80
    
    # Save with lock detection
    saved_path = save_excel_safely(wb, output_file)
    
    print(f"\n{'='*80}")
    print("SUCCESS - ALL ITEMS EXTRACTED FROM IMAGES")
    print(f"{'='*80}\n")
    print(f"Output: {saved_path.absolute()}")
    print(f"Total items: {len(valid_items)}")
    print(f"Work Order: Rs. {total:,.2f}")
    print(f"Bill Amount: Rs. {bill_total:,.2f}")
    
    print(f"\nItems extracted:")
    for item in valid_items[:10]:  # Show first 10
        print(f"   {item['code']:10} | {item['unit']:8} | Rs.{item['rate']:8,.2f} | {item['description'][:50]}")
    
    if len(valid_items) > 10:
        print(f"   ... and {len(valid_items) - 10} more items")
    
    return 0

if __name__ == '__main__':
    sys.exit(main())