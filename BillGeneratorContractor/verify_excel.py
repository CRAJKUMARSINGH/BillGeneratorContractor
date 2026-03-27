#!/usr/bin/env python3
"""Verify Excel file structure"""
import openpyxl

wb = openpyxl.load_workbook('OUTPUT/INPUT_work_01_27022026_PRODUCTION.xlsx')
print('Sheets:', wb.sheetnames)

print('\n' + '='*80)
print('WORK ORDER SHEET:')
print('='*80)
ws_wo = wb['Work Order']
print(f'Total Rows: {ws_wo.max_row}')
for i in range(1, min(8, ws_wo.max_row+1)):
    item = ws_wo.cell(i,1).value
    desc = ws_wo.cell(i,2).value
    qty = ws_wo.cell(i,4).value
    print(f'Row {i}: Item={item}, Qty={qty}, Desc={desc[:40] if desc else None}...')

print('\n' + '='*80)
print('BILL QUANTITY SHEET:')
print('='*80)
ws_bq = wb['Bill Quantity']
print(f'Total Rows: {ws_bq.max_row}')
for i in range(1, min(8, ws_bq.max_row+1)):
    item = ws_bq.cell(i,1).value
    desc = ws_bq.cell(i,2).value
    qty = ws_bq.cell(i,4).value
    print(f'Row {i}: Item={item}, Qty={qty}, Desc={desc[:40] if desc else None}...')

print('\n' + '='*80)
print('COMPARISON:')
print('='*80)
print(f'Work Order rows: {ws_wo.max_row}')
print(f'Bill Quantity rows: {ws_bq.max_row}')
print(f'Are they the same? {ws_wo.max_row == ws_bq.max_row}')

# Check if quantities match
print('\nQuantity Comparison:')
for i in range(2, min(8, ws_wo.max_row+1)):
    wo_qty = ws_wo.cell(i,4).value
    bq_qty = ws_bq.cell(i,4).value
    match = '✓' if wo_qty == bq_qty else '✗'
    print(f'Row {i}: WO={wo_qty}, BQ={bq_qty} {match}')
