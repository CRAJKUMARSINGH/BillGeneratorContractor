#!/usr/bin/env python3
"""
Complete Bill Documentation Generator
End-to-end pipeline: Images → Excel → All Bill PDFs

This script:
1. Extracts data from work order images using FINAL extraction system
2. Generates Excel input file with quantities
3. Creates all bill documents as PDFs:
   - First Page Bill
   - Deviation Statement
   - Certificates (Material, Labour, Measurement)
   - Abstract of Cost
   - All other required documents

Usage:
    python generate_complete_bill_docs.py
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from modules.multi_layer_extractor import MultiLayerExtractor
from modules.confidence_scorer import ConfidenceScorer
from modules.pwd_database import PWDDatabase
from modules.api_key_manager import APIKeyManager, APIKey
from modules.image_quality_checker import ImageQualityChecker
from modules.image_preprocessor import ImagePreprocessor
from modules.completeness_checker import CompletenessChecker

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import traceback
import time
import json

# Configuration
INPUT_FOLDER = Path("INPUT_WORK_ORDER_IMAGES_TEXT")
QTY_FILE = INPUT_FOLDER / "qty.txt"
OUTPUT_FOLDER = Path("OUTPUT")
EXCEL_OUTPUT = OUTPUT_FOLDER / "BILL_INPUT_COMPLETE.xlsx"
LOG_FILE = OUTPUT_FOLDER / "complete_bill_generation_log.txt"

# Quality thresholds
MIN_QUALITY_SCORE = 0.5
MIN_CONFIDENCE = 0.7

# API Keys
API_KEYS = [
    APIKey(key="AIzaSyBMZYPgjcqXY-tpe6UhtBtrWhzfbU0-WVU", name="Primary", daily_quota=20),
    APIKey(key="AIzaSyDCU_qa6mH4Dz0Rcvof7RQrr8P6HevZJpc", name="Backup1", daily_quota=20),
]

def log_message(message: str, level: str = "INFO"):
    """Log message to file and console"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] [{level}] {message}"
    print(log_entry)
    
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(log_entry + "\n")

def read_qty_file(qty_file_path):
    """Read quantities from qty.txt"""
    qty_data = {}
    try:
        with open(qty_file_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split()
                    if len(parts) >= 2:
                        item_no = parts[0]
                        quantity = float(parts[1])
                        qty_data[item_no] = quantity
        log_message(f"Loaded {len(qty_data)} quantities from {qty_file_path.name}")
    except Exception as e:
        log_message(f"Error reading qty file: {e}", "ERROR")
    return qty_data

def sort_items_by_bsr(items):
    """Sort items by BSR code numerically"""
    def sort_key(item):
        code = item.get('code', '0')
        parts = code.split('.')
        try:
            return tuple(int(p) for p in parts)
        except:
            return (999, 999, 999)
    
    return sorted(items, key=sort_key)

def extract_from_images(image_folder, db, key_manager):
    """
    Extract all items from work order images
    Returns: list of extracted items
    """
    log_message("\n" + "="*80)
    log_message("STEP 1: EXTRACTING DATA FROM IMAGES")
    log_message("="*80)
    
    # Initialize components
    scorer = ConfidenceScorer(db)
    quality_checker = ImageQualityChecker()
    preprocessor = ImagePreprocessor()
    
    current_key = key_manager.get_current_key()
    if not current_key:
        log_message("No available API keys!", "ERROR")
        return []
    
    extractor = MultiLayerExtractor(gemini_api_key=current_key.key)
    
    # Get images
    image_files = sorted(image_folder.glob("*.jpg")) + sorted(image_folder.glob("*.jpeg"))
    log_message(f"Found {len(image_files)} images to process")
    
    all_items = []
    stats = {'success': 0, 'failed': 0}
    
    for i, image_path in enumerate(image_files, 1):
        log_message(f"\n[{i}/{len(image_files)}] Processing {image_path.name}")
        
        try:
            # Quality check
            quality_result = quality_checker.check_quality(str(image_path))
            log_message(f"  Quality: {quality_result.score:.2f}")
            
            if quality_result.score < MIN_QUALITY_SCORE:
                log_message(f"  Preprocessing image...", "WARNING")
                processed_path = preprocessor.preprocess(str(image_path))
            else:
                processed_path = str(image_path)
            
            # Extract with retry
            for attempt in range(1, 4):
                try:
                    result = extractor.extract_with_fallback(processed_path, min_confidence=MIN_CONFIDENCE)
                    
                    if result.success and result.items:
                        all_items.extend(result.items)
                        stats['success'] += 1
                        log_message(f"  Extracted {len(result.items)} items")
                        break
                    
                    if attempt < 3:
                        time.sleep(2 ** (attempt - 1))
                
                except Exception as e:
                    if attempt == 3:
                        stats['failed'] += 1
                        log_message(f"  Extraction failed: {e}", "ERROR")
        
        except Exception as e:
            stats['failed'] += 1
            log_message(f"  Error: {e}", "ERROR")
    
    # Remove duplicates
    unique_items = {}
    for item in all_items:
        code = item.get('code', '')
        if code and code not in unique_items:
            unique_items[code] = item
    
    items_list = list(unique_items.values())
    items_list = sort_items_by_bsr(items_list)
    
    log_message(f"\nExtraction complete: {len(items_list)} unique items")
    log_message(f"Success rate: {(stats['success']/(stats['success']+stats['failed'])*100):.1f}%")
    
    return items_list

def create_excel_input(items, qty_data, output_file, db):
    """
    Create Excel input file with all extracted data
    """
    log_message("\n" + "="*80)
    log_message("STEP 2: CREATING EXCEL INPUT FILE")
    log_message("="*80)
    
    scorer = ConfidenceScorer(db)
    scores = scorer.score_items(items)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Create sheets
    ws_work = wb.create_sheet("Work Order")
    ws_bill = wb.create_sheet("Bill Quantity")
    ws_summary = wb.create_sheet("Summary")
    
    # Headers
    headers = ["Item", "Description", "Unit", "Quantity", "Rate", "Amount", "BSR", "Confidence"]
    
    # Work Order sheet
    ws_work.append(headers)
    for cell in ws_work[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for i, (item, score) in enumerate(zip(items, scores), 1):
        code = item.get('code', '')
        desc = item.get('description', '')
        unit = item.get('unit', '')
        qty = item.get('quantity', 1)
        rate = item.get('rate', 0)
        amount = qty * rate
        
        row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}"]
        ws_work.append(row)
    
    # Bill Quantity sheet
    ws_bill.append(headers)
    for cell in ws_bill[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for i, (item, score) in enumerate(zip(items, scores), 1):
        code = item.get('code', '')
        desc = item.get('description', '')
        unit = item.get('unit', '')
        rate = item.get('rate', 0)
        qty = qty_data.get(code, 0)
        amount = qty * rate
        
        row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}"]
        ws_bill.append(row)
    
    # Summary sheet
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    summary_data = [
        ["Complete Bill Documentation", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Items", len(items)],
        ["Work Order Items", len(items)],
        ["Bill Quantity Items", len([q for q in qty_data.values() if q > 0])],
        ["", ""],
        ["Average Confidence", f"{sum(s.overall for s in scores)/len(scores):.2f}"],
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Save
    output_file.parent.mkdir(exist_ok=True)
    wb.save(output_file)
    
    log_message(f"Excel file saved: {output_file}")
    return output_file

def generate_bill_pdfs(excel_file, output_folder):
    """
    Generate all bill PDFs from Excel input
    """
    log_message("\n" + "="*80)
    log_message("STEP 3: GENERATING BILL PDFs")
    log_message("="*80)
    
    try:
        # Import PDF generator
        from core.generators.pdf_generator_fixed import FixedPDFGenerator
        from core.config.config_loader import ConfigLoader
        
        # Load config
        config = ConfigLoader.load_from_env('BILL_CONFIG', 'config/v01.json')
        
        # Initialize generator
        generator = FixedPDFGenerator(config)
        
        # Read Excel data
        import pandas as pd
        df = pd.read_excel(excel_file, sheet_name="Bill Quantity")
        
        # Generate all documents
        documents = [
            ("First Page Bill", "first_page"),
            ("Deviation Statement", "deviation"),
            ("Material Certificate", "material_cert"),
            ("Labour Certificate", "labour_cert"),
            ("Measurement Certificate", "measurement_cert"),
            ("Abstract of Cost", "abstract"),
        ]
        
        generated_files = []
        
        for doc_name, doc_type in documents:
            try:
                log_message(f"Generating {doc_name}...")
                
                # Generate PDF
                pdf_path = output_folder / f"{doc_type.upper()}.pdf"
                
                # Use appropriate template
                if doc_type == "first_page":
                    html_content = generator.generate_first_page_html(df)
                elif doc_type == "deviation":
                    html_content = generator.generate_deviation_html(df)
                elif doc_type == "material_cert":
                    html_content = generator.generate_certificate_html("Material", df)
                elif doc_type == "labour_cert":
                    html_content = generator.generate_certificate_html("Labour", df)
                elif doc_type == "measurement_cert":
                    html_content = generator.generate_certificate_html("Measurement", df)
                elif doc_type == "abstract":
                    html_content = generator.generate_abstract_html(df)
                
                # Convert to PDF
                pdf_bytes = generator.convert_html_to_pdf(html_content, landscape=True)
                
                # Save
                with open(pdf_path, 'wb') as f:
                    f.write(pdf_bytes)
                
                generated_files.append(pdf_path)
                log_message(f"  ✓ {doc_name} saved: {pdf_path.name}")
            
            except Exception as e:
                log_message(f"  ✗ Failed to generate {doc_name}: {e}", "ERROR")
        
        log_message(f"\nGenerated {len(generated_files)} PDF documents")
        return generated_files
    
    except Exception as e:
        log_message(f"PDF generation error: {e}", "ERROR")
        log_message(traceback.format_exc(), "DEBUG")
        return []

def main():
    """Main pipeline"""
    log_message("="*80)
    log_message("COMPLETE BILL DOCUMENTATION GENERATOR")
    log_message("Images → Excel → All Bill PDFs")
    log_message("="*80)
    
    start_time = time.time()
    
    try:
        # Initialize
        log_message("\nInitializing components...")
        db = PWDDatabase()
        key_manager = APIKeyManager(API_KEYS)
        
        log_message(f"PWD Database: {len(db.items)} items")
        log_message(f"API Keys: {key_manager.get_status()['active_keys']} active")
        
        # Step 1: Extract from images
        items = extract_from_images(INPUT_FOLDER, db, key_manager)
        
        if not items:
            log_message("No items extracted. Exiting.", "ERROR")
            return
        
        # Step 2: Read quantities
        qty_data = read_qty_file(QTY_FILE)
        
        # Step 3: Create Excel
        excel_file = create_excel_input(items, qty_data, EXCEL_OUTPUT, db)
        
        # Step 4: Generate PDFs
        pdf_files = generate_bill_pdfs(excel_file, OUTPUT_FOLDER)
        
        # Final report
        elapsed_time = time.time() - start_time
        
        log_message("\n" + "="*80)
        log_message("COMPLETE BILL GENERATION SUCCESSFUL")
        log_message("="*80)
        log_message(f"Processing time: {elapsed_time:.1f}s")
        log_message(f"Items extracted: {len(items)}")
        log_message(f"Excel file: {excel_file}")
        log_message(f"PDF documents: {len(pdf_files)}")
        log_message("\nGenerated files:")
        log_message(f"  📊 {excel_file}")
        for pdf in pdf_files:
            log_message(f"  📄 {pdf}")
        log_message("="*80)
    
    except Exception as e:
        log_message(f"\nFATAL ERROR: {e}", "ERROR")
        log_message(traceback.format_exc(), "DEBUG")
        raise

if __name__ == '__main__':
    main()
