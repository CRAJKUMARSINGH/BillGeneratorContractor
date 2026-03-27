#!/usr/bin/env python3
"""
Production-Ready Extraction - Week 4 Complete
Multi-layer + Validation + Retry + Error Handling
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from modules.multi_layer_extractor import MultiLayerExtractor
from modules.confidence_scorer import ConfidenceScorer
from modules.pwd_database import PWDDatabase
from modules.retry_handler import retry, STANDARD_RETRY, RetryHandler
from modules.api_key_manager import APIKeyManager, APIKey
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import traceback

# Configuration
INPUT_FOLDER = Path("INPUT_WORK_ORDER_IMAGES_TEXT")
QTY_FILE = INPUT_FOLDER / "qty.txt"
OUTPUT_FILE = Path("OUTPUT/INPUT_PRODUCTION_READY.xlsx")
LOG_FILE = Path("OUTPUT/extraction_log.txt")

# API Keys (with rotation)
API_KEYS = [
    APIKey(key="AIzaSyBMZYPgjcqXY-tpe6UhtBtrWhzfbU0-WVU", name="Primary", daily_quota=20),
    APIKey(key="AIzaSyDCU_qa6mH4Dz0Rcvof7RQrr8P6HevZJpc", name="Backup1", daily_quota=20),
]

def log_message(message: str, level: str = "INFO"):
    """Log message to file and console"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] [{level}] {message}"
    print(log_entry)
    
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(log_entry + "\n")

def read_qty_file(qty_file_path):
    """Read quantities from qty.txt"""
    qty_data = {}
    try:
        with open(qty_file_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split()
                    if len(parts) >= 2:
                        item_no = parts[0]
                        quantity = float(parts[1])
                        qty_data[item_no] = quantity
        log_message(f"Loaded {len(qty_data)} quantities from {qty_file_path.name}")
    except Exception as e:
        log_message(f"Error reading qty file: {e}", "ERROR")
    return qty_data

def sort_items_by_bsr(items):
    """Sort items by BSR code numerically"""
    def sort_key(item):
        code = item.get('code', '0')
        parts = code.split('.')
        try:
            return tuple(int(p) for p in parts)
        except:
            return (999, 999, 999)
    
    return sorted(items, key=sort_key)

def extract_with_retry_and_rotation(extractor, image_path, key_manager, max_attempts=3):
    """Extract with retry and API key rotation"""
    for attempt in range(1, max_attempts + 1):
        try:
            # Get current key
            current_key = key_manager.get_current_key()
            if not current_key:
                log_message("No available API keys", "ERROR")
                return None
            
            log_message(f"Attempt {attempt}/{max_attempts} using {current_key.name}")
            
            # Try extraction
            result = extractor.extract_with_fallback(str(image_path), min_confidence=0.7)
            
            if result.success:
                key_manager.mark_current_used()
                return result
            
            # Check error type
            if result.error and "429" in result.error:
                # Quota exceeded
                log_message(f"Quota exceeded for {current_key.name}", "WARNING")
                key_manager.mark_current_quota_exceeded()
            elif result.error and ("401" in result.error or "403" in result.error):
                # Invalid key
                log_message(f"Invalid key: {current_key.name}", "ERROR")
                key_manager.mark_current_invalid(result.error)
            else:
                # Other error - retry with same key
                if attempt < max_attempts:
                    import time
                    delay = 2 ** (attempt - 1)
                    log_message(f"Retrying in {delay}s...", "WARNING")
                    time.sleep(delay)
        
        except Exception as e:
            log_message(f"Extraction error: {e}", "ERROR")
            if attempt < max_attempts:
                import time
                time.sleep(2 ** (attempt - 1))
    
    return None

def main():
    log_message("="*80)
    log_message("PRODUCTION-READY EXTRACTION - WEEK 4 COMPLETE")
    log_message("="*80)
    
    try:
        # Initialize components
        log_message("Initializing components...")
        
        key_manager = APIKeyManager(API_KEYS)
        current_key = key_manager.get_current_key()
        
        if not current_key:
            log_message("No available API keys!", "ERROR")
            return
        
        extractor = MultiLayerExtractor(gemini_api_key=current_key.key)
        db = PWDDatabase()
        scorer = ConfidenceScorer(db)
        
        # Show status
        key_status = key_manager.get_status()
        log_message(f"API Keys: {key_status['active_keys']}/{key_status['total_keys']} active")
        
        extractor_status = extractor.get_status()
        log_message(f"Extraction Layers: {extractor_status['available_extractors']}/{extractor_status['total_extractors']} available")
        
        # Get images
        image_files = sorted(INPUT_FOLDER.glob("*.jpg")) + sorted(INPUT_FOLDER.glob("*.jpeg"))
        log_message(f"Found {len(image_files)} images to process")
        
        # Extract with retry and rotation
        log_message("="*80)
        log_message("EXTRACTING WITH RETRY & KEY ROTATION")
        log_message("="*80)
        
        all_items = []
        extraction_stats = {
            'success': 0,
            'failed': 0,
            'retries': 0
        }
        
        for i, image_path in enumerate(image_files, 1):
            log_message(f"\n[{i}/{len(image_files)}] {image_path.name}")
            log_message("-" * 80)
            
            try:
                result = extract_with_retry_and_rotation(
                    extractor, image_path, key_manager, max_attempts=3
                )
                
                if result and result.success and result.items:
                    all_items.extend(result.items)
                    extraction_stats['success'] += 1
                    log_message(f"SUCCESS: Extracted {len(result.items)} items")
                else:
                    extraction_stats['failed'] += 1
                    log_message(f"FAILED: No items extracted", "ERROR")
            
            except Exception as e:
                extraction_stats['failed'] += 1
                log_message(f"ERROR: {e}", "ERROR")
                log_message(traceback.format_exc(), "DEBUG")
        
        # Remove duplicates
        unique_items = {}
        for item in all_items:
            code = item.get('code', '')
            if code and code not in unique_items:
                unique_items[code] = item
        
        items_list = list(unique_items.values())
        
        log_message("="*80)
        log_message("EXTRACTION COMPLETE")
        log_message("="*80)
        log_message(f"Total items: {len(all_items)}")
        log_message(f"Unique items: {len(items_list)}")
        log_message(f"Success rate: {(extraction_stats['success']/len(image_files)*100):.1f}%")
        
        # Sort and validate
        items_list = sort_items_by_bsr(items_list)
        
        log_message("="*80)
        log_message("VALIDATING ITEMS")
        log_message("="*80)
        
        scores = scorer.score_items(items_list)
        report = scorer.generate_report(items_list)
        
        log_message(f"Average Confidence: {report['average_confidence']:.2f}")
        log_message(f"Auto-Accept: {report['recommended_actions']['auto_accept']['percentage']:.1f}%")
        
        # Read quantities
        qty_data = read_qty_file(QTY_FILE)
        
        # Create Excel
        log_message("="*80)
        log_message("CREATING EXCEL")
        log_message("="*80)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["PRODUCTION-READY EXTRACTION REPORT"])
        ws_summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_summary.append([])
        ws_summary.append(["EXTRACTION STATISTICS"])
        ws_summary.append(["Total Images:", len(image_files)])
        ws_summary.append(["Successful:", extraction_stats['success']])
        ws_summary.append(["Failed:", extraction_stats['failed']])
        ws_summary.append(["Success Rate:", f"{(extraction_stats['success']/len(image_files)*100):.1f}%"])
        ws_summary.append([])
        ws_summary.append(["VALIDATION STATISTICS"])
        ws_summary.append(["Total Items:", len(items_list)])
        ws_summary.append(["Average Confidence:", f"{report['average_confidence']:.2f}"])
        ws_summary.append(["Auto-Accept:", f"{report['recommended_actions']['auto_accept']['percentage']:.1f}%"])
        
        # Work Order sheet
        ws_work = wb.create_sheet("Work Order")
        headers = ["Item", "Description", "Unit", "Quantity", "Rate", "Amount", "BSR", "Confidence", "Status"]
        ws_work.append(headers)
        
        for cell in ws_work[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        for i, (item, score) in enumerate(zip(items_list, scores), 1):
            code = item.get('code', '')
            desc = item.get('description', '')
            unit = item.get('unit', '')
            qty = item.get('quantity', 1)
            rate = item.get('rate', 0)
            amount = qty * rate
            
            if score.overall >= 0.95:
                status, fill_color = "AUTO_ACCEPT", "C6EFCE"
            elif score.overall >= 0.85:
                status, fill_color = "QUICK_REVIEW", "FFEB9C"
            elif score.overall >= 0.70:
                status, fill_color = "REVIEW", "FFC7CE"
            else:
                status, fill_color = "DETAILED_REVIEW", "FF6B6B"
            
            row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}", status]
            ws_work.append(row)
            
            for cell in ws_work[i+1]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Bill Quantity sheet
        ws_bill = wb.create_sheet("Bill Quantity")
        ws_bill.append(headers)
        
        for cell in ws_bill[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        for i, (item, score) in enumerate(zip(items_list, scores), 1):
            code = item.get('code', '')
            desc = item.get('description', '')
            unit = item.get('unit', '')
            rate = item.get('rate', 0)
            qty = qty_data.get(code, 0)
            amount = qty * rate
            
            if score.overall >= 0.95:
                status, fill_color = "AUTO_ACCEPT", "C6EFCE"
            elif score.overall >= 0.85:
                status, fill_color = "QUICK_REVIEW", "FFEB9C"
            elif score.overall >= 0.70:
                status, fill_color = "REVIEW", "FFC7CE"
            else:
                status, fill_color = "DETAILED_REVIEW", "FF6B6B"
            
            row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}", status]
            ws_bill.append(row)
            
            for cell in ws_bill[i+1]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Save
        OUTPUT_FILE.parent.mkdir(exist_ok=True)
        wb.save(OUTPUT_FILE)
        
        log_message(f"Excel saved: {OUTPUT_FILE}")
        
        # Final statistics
        log_message("="*80)
        log_message("WEEK 4 COMPLETE - PRODUCTION READY")
        log_message("="*80)
        log_message(f"Reliability: {(extraction_stats['success']/len(image_files)*100):.1f}%")
        log_message(f"Validation Confidence: {report['average_confidence']*100:.1f}%")
        log_message(f"Auto-Accept Rate: {report['recommended_actions']['auto_accept']['percentage']:.1f}%")
        log_message("="*80)
        
    except Exception as e:
        log_message(f"FATAL ERROR: {e}", "ERROR")
        log_message(traceback.format_exc(), "DEBUG")
        raise

if __name__ == '__main__':
    main()
