#!/usr/bin/env python3
"""
Create INPUT Excel from Gemini extracted data
Uses ONLY data from images - NO database needed
"""
import sys
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

def load_gemini_data(json_file):
    """Load extracted data from Gemini"""
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

def load_qty_data(qty_file):
    """Load quantities from qty.txt"""
    qty_data = {}
    if Path(qty_file).exists():
        with open(qty_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split()
                    if len(parts) >= 2:
                        code = parts[0]
                        qty = float(parts[1])
                        qty_data[code] = qty
    return qty_data

def parse_gemini_json(text):
    """Parse JSON from Gemini response"""
    try:
        # Try direct JSON parse
        return json.loads(text)
    except:
        # Try to extract JSON from markdown code blocks
        import re
        json_match = re.search(r'```json\s*(\{.*?\})\s*```', text, re.DOTALL)
        if json_match:
            return json.loads(json_match.group(1))
        
        # Try to find JSON object
        json_match = re.search(r'\{.*\}', text, re.DOTALL)
        if json_match:
            return json.loads(json_match.group(0))
        
        return None

def create_input_excel(gemini_data, qty_data, output_file):
    """Create INPUT Excel from Gemini data"""
    
    print(f"\n{'='*80}")
    print("📊 CREATING INPUT EXCEL FROM GEMINI DATA")
    print(f"{'='*80}\n")
    
    # Combine all items from all images
    all_items = []
    header_info = {}
    
    for result in gemini_data:
        print(f"📄 Processing: {result['image']}")
        
        # Parse JSON
        parsed = parse_gemini_json(result['data'])
        
        if not parsed:
            print(f"   ⚠️ Could not parse JSON")
            continue
        
        # Extract items
        if 'items' in parsed:
            items = parsed['items']
            print(f"   ✅ Found {len(items)} items")
            all_items.extend(items)
        
        # Extract header info
        if 'header_info' in parsed:
            header_info.update(parsed['header_info'])
    
    print(f"\n✅ Total items extracted: {len(all_items)}")
    
    # Create Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # === SHEET 1: Title ===
    ws_title = wb.create_sheet("Title")
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED", ""],
        ["Bill Number", "First"],
        ["Running or Final", "Final"],
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", header_info.get('contractor_name', '[FROM IMAGES]')],
        ["Name of Work ;-", header_info.get('work_name', '[FROM IMAGES]')],
        ["Serial No. of this bill :", "First & Final Bill"],
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", header_info.get('work_order_number', '[FROM IMAGES]')],
        ["Agreement No.", header_info.get('agreement_number', '[FROM IMAGES]')],
        ["WORK ORDER AMOUNT RS.", header_info.get('total_amount', '[FROM IMAGES]')],
        ["Date of written order to commence work :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of Start :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of completion :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of actual completion of work :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of measurement :", datetime.now().strftime("%Y-%m-%d")],
        ["TENDER PREMIUM %", "11.22"],
        ["Above / Below", "Above"],
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        cell_label = ws_title.cell(row=row_idx, column=1, value=label)
        cell_value = ws_title.cell(row=row_idx, column=2, value=value)
        cell_label.font = Font(bold=True)
    
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # === SHEET 2: Work Order ===
    ws_work = wb.create_sheet("Work Order")
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_work.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    total_wo_amount = 0.0
    
    for item in all_items:
        code = item.get('code', item.get('bsr_code', ''))
        desc = item.get('description', '')
        unit = item.get('unit', '')
        rate = float(item.get('rate', 0))
        qty = float(item.get('quantity', 0))
        
        # If no quantity in item, try to get from qty_data
        if qty == 0 and code in qty_data:
            qty = qty_data[code]
        
        amount = qty * rate if qty > 0 else 0
        
        ws_work.cell(row=row_idx, column=1, value=code)
        ws_work.cell(row=row_idx, column=2, value=desc)
        ws_work.cell(row=row_idx, column=3, value=unit)
        ws_work.cell(row=row_idx, column=4, value=qty if qty > 0 else '')
        ws_work.cell(row=row_idx, column=5, value=rate)
        ws_work.cell(row=row_idx, column=6, value=amount if amount > 0 else '')
        ws_work.cell(row=row_idx, column=7, value=code)
        
        if amount > 0:
            total_wo_amount += amount
        
        row_idx += 1
    
    ws_work.column_dimensions['B'].width = 80
    
    # === SHEET 3: Bill Quantity ===
    ws_bill = wb.create_sheet("Bill Quantity")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bill.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    for item in all_items:
        code = item.get('code', item.get('bsr_code', ''))
        desc = item.get('description', '')
        unit = item.get('unit', '')
        rate = float(item.get('rate', 0))
        
        # Get quantity from qty.txt
        qty = qty_data.get(code, 0)
        amount = qty * rate if qty > 0 else 0
        
        ws_bill.cell(row=row_idx, column=1, value=code)
        ws_bill.cell(row=row_idx, column=2, value=desc)
        ws_bill.cell(row=row_idx, column=3, value=unit)
        ws_bill.cell(row=row_idx, column=4, value=qty if qty > 0 else '')
        ws_bill.cell(row=row_idx, column=5, value=rate)
        ws_bill.cell(row=row_idx, column=6, value=amount if amount > 0 else '')
        ws_bill.cell(row=row_idx, column=7, value=code)
        
        row_idx += 1
    
    ws_bill.column_dimensions['B'].width = 80
    
    # === SHEET 4: Extra Items ===
    ws_extra = wb.create_sheet("Extra Items")
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        cell = ws_extra.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    ws_extra.column_dimensions['B'].width = 80
    
    # Save Excel
    wb.save(output_file)
    
    print(f"\n{'='*80}")
    print("✅ INPUT EXCEL CREATED FROM GEMINI DATA")
    print(f"{'='*80}\n")
    print(f"📁 Output: {output_file}")
    print(f"📊 Total items: {len(all_items)}")
    print(f"💰 Work Order Amount: Rs. {total_wo_amount:,.2f}")
    print(f"\n🎯 Next: python process_first_bill.py {output_file}")
    
    return True

def main():
    gemini_json = "OUTPUT/GEMINI_EXTRACTED_DATA.json"
    qty_file = "INPUT/work_order_samples/work_01_27022026/qty.txt"
    output_file = "OUTPUT/INPUT_FROM_GEMINI.xlsx"
    
    if len(sys.argv) > 1:
        gemini_json = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    # Load data
    print("📄 Loading Gemini extracted data...")
    gemini_data = load_gemini_data(gemini_json)
    
    print("📄 Loading quantities from qty.txt...")
    qty_data = load_qty_data(qty_file)
    
    # Create Excel
    success = create_input_excel(gemini_data, qty_data, output_file)
    
    return 0 if success else 1

if __name__ == '__main__':
    sys.exit(main())
