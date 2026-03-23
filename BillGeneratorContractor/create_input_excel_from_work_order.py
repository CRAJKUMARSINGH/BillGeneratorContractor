#!/usr/bin/env python3
"""
STEP 1: Create INPUT Excel file from Work Order images and QTY.txt
Matches TEST_INPUT_FILES format exactly
"""
import sys
from pathlib import Path
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

# Add project root
sys.path.insert(0, str(Path(__file__).parent))

def read_qty_file(qty_file_path):
    """Read QTY.txt file and parse item quantities"""
    qty_data = {}
    
    with open(qty_file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                parts = line.split()
                if len(parts) >= 2:
                    item_no = parts[0]
                    quantity = float(parts[1])
                    qty_data[item_no] = quantity
    
    return qty_data

def create_input_excel(work_order_dir, output_file):
    """
    Create Excel file matching TEST_INPUT_FILES format
    Based on work order images and QTY.txt
    """
    
    print(f"\n{'='*80}")
    print(f"STEP 1: Creating INPUT Excel File")
    print(f"{'='*80}\n")
    
    work_order_path = Path(work_order_dir)
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Read QTY.txt
    qty_file = work_order_path / "qty.txt"
    if not qty_file.exists():
        print(f"❌ QTY.txt not found at: {qty_file}")
        return False
    
    print(f"Reading QTY data from: {qty_file}")
    qty_data = read_qty_file(qty_file)
    print(f"✅ Found {len(qty_data)} items with quantities:")
    for item_no, qty in qty_data.items():
        print(f"   {item_no}: {qty}")
    print()
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # ========================================================================
    # SHEET 1: Title (Matching FirstFINALnoExtra.xlsx exactly)
    # ========================================================================
    print("Creating Title sheet...")
    ws_title = wb.create_sheet("Title")
    
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED", ""],
        ["Bill Number", "First"],
        ["Running or Final", "Final"],
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", "M/s. [CONTRACTOR NAME FROM IMAGES]"],
        ["Name of Work ;-", "[WORK NAME FROM IMAGES]"],
        ["Serial No. of this bill :", "First & Final Bill"],
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", "[WO NUMBER FROM IMAGES]"],
        ["Agreement No.", "[AGREEMENT NO FROM IMAGES]"],
        ["WORK ORDER AMOUNT RS.", "[AMOUNT FROM IMAGES]"],
        ["Date of written order to commence work :", "2026-02-27"],
        ["St. date of Start :", "2026-03-01"],
        ["St. date of completion :", "2026-06-30"],
        ["Date of actual completion of work :", "2026-06-30"],
        ["Date of measurement :", "2026-03-09"],
        ["TENDER PREMIUM %", "11.22"],
        ["Above / Below", "Above"],
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        ws_title.cell(row=row_idx, column=1, value=label)
        ws_title.cell(row=row_idx, column=2, value=value)
    
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # ========================================================================
    # SHEET 2: Work Order (Based on images - MANUAL ENTRY REQUIRED)
    # ========================================================================
    print("Creating Work Order sheet...")
    ws_work = wb.create_sheet("Work Order")
    
    # Headers
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_work.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Sample data structure based on QTY.txt items
    # YOU NEED TO FILL DESCRIPTIONS, UNITS, AND RATES FROM IMAGES
    work_order_items = [
        # Main item 1
        [1.0, "Rewiring of light point/ fan point/ exhaust fan point [FULL DESCRIPTION FROM IMAGE]", None, None, None, None, "1.5"],
        # Sub-items from QTY.txt
        [None, "Short point (up to 3 mtr.)", "P. point", 50.0, 256, 12800.0, "1.5.1"],
        [None, "Medium point (up to 6 mtr.)", "P. point", 50.0, 472, 23600.0, "1.5.2"],
        [None, "Long point (up to 10 mtr.)", "P. point", 50.0, 662, 33100.0, "1.5.3"],
        
        # Main item 2
        [2.0, "Rewiring of 3/5 pin 6 amp. Light plug point [FULL DESCRIPTION FROM IMAGE]", None, None, None, None, "1.7"],
        [None, "On board", "P. point", 100.0, 136, 13600.0, "1.7.1"],
        
        # Item 3
        [3.0, "P & F ISI marked (IS:3854) 6 amp. flush type non modular switch [FULL DESCRIPTION FROM IMAGE]", "Each", 10.0, 23, 230.0, "7.1"],
        
        # Item 4
        [4.0, "P & F ISI marked (IS:3854) 16 amp. flush type non modular switch [FULL DESCRIPTION FROM IMAGE]", "Each", 30.0, 50, 1500.0, "7.2"],
        
        # Item 5
        [5.0, "Providing & Fixing of 3/5 pin 6 amp. flush type non modular socket [FULL DESCRIPTION FROM IMAGE]", "Each", 10.0, 33, 330.0, "7.10"],
        
        # Item 6
        [6.0, "Providing & Fixing of 3/6 pin 16 amp flush type non modular socket [FULL DESCRIPTION FROM IMAGE]", "Each", 10.0, 78, 780.0, "7.11"],
        
        # Add more items as per your images...
    ]
    
    for row_idx, row_data in enumerate(work_order_items, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_work.cell(row=row_idx, column=col_idx, value=value)
    
    ws_work.column_dimensions['A'].width = 8
    ws_work.column_dimensions['B'].width = 80
    ws_work.column_dimensions['C'].width = 12
    ws_work.column_dimensions['D'].width = 12
    ws_work.column_dimensions['E'].width = 12
    ws_work.column_dimensions['F'].width = 15
    ws_work.column_dimensions['G'].width = 10
    
    # ========================================================================
    # SHEET 3: Bill Quantity (Apply QTY.txt data)
    # ========================================================================
    print("Creating Bill Quantity sheet with QTY data...")
    ws_bill = wb.create_sheet("Bill Quantity")
    
    # Copy headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bill.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Copy work order items and update quantities from QTY.txt
    # Items NOT in QTY.txt have ZERO quantity
    bill_items = []
    for item in work_order_items:
        bill_item = item.copy()
        
        # Check if this item has quantity in QTY.txt
        bsr_code = item[6]  # BSR column
        if bsr_code and bsr_code in qty_data:
            # Update quantity from QTY.txt
            new_qty = qty_data[bsr_code]
            bill_item[3] = new_qty  # Quantity column
            
            # Recalculate amount if rate exists
            if bill_item[4]:  # Rate exists
                bill_item[5] = new_qty * bill_item[4]  # Amount = Qty × Rate
            
            print(f"   Updated {bsr_code}: Quantity = {new_qty}, Amount = {bill_item[5]}")
        else:
            # Item NOT in QTY.txt = ZERO quantity
            if bsr_code and bill_item[3] is not None:  # Has a quantity in work order
                bill_item[3] = 0.0  # Set to ZERO
                bill_item[5] = 0.0  # Amount also ZERO
                print(f"   {bsr_code}: Quantity = 0 (not in QTY.txt)")
        
        bill_items.append(bill_item)
    
    for row_idx, row_data in enumerate(bill_items, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_bill.cell(row=row_idx, column=col_idx, value=value)
    
    ws_bill.column_dimensions['A'].width = 8
    ws_bill.column_dimensions['B'].width = 80
    ws_bill.column_dimensions['C'].width = 12
    ws_bill.column_dimensions['D'].width = 12
    ws_bill.column_dimensions['E'].width = 12
    ws_bill.column_dimensions['F'].width = 15
    ws_bill.column_dimensions['G'].width = 10
    
    # ========================================================================
    # SHEET 4: Extra Items (Empty - no extra items)
    # ========================================================================
    print("Creating Extra Items sheet (empty)...")
    ws_extra = wb.create_sheet("Extra Items")
    
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        cell = ws_extra.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    ws_extra.column_dimensions['A'].width = 8
    ws_extra.column_dimensions['B'].width = 80
    ws_extra.column_dimensions['C'].width = 12
    ws_extra.column_dimensions['D'].width = 12
    ws_extra.column_dimensions['E'].width = 12
    ws_extra.column_dimensions['F'].width = 15
    ws_extra.column_dimensions['G'].width = 12
    ws_extra.column_dimensions['H'].width = 10
    
    # Save workbook
    wb.save(output_path)
    
    print(f"\n✅ Excel file created: {output_path.absolute()}")
    print(f"\n{'='*80}")
    print("IMPORTANT: Manual Steps Required")
    print(f"{'='*80}")
    print("\n1. Open the Excel file:")
    print(f"   {output_path.absolute()}")
    print("\n2. Open the work order images:")
    print(f"   {work_order_path.absolute()}")
    print("\n3. Fill in the following from images:")
    print("   Title Sheet:")
    print("     - Row 5: Contractor Name")
    print("     - Row 6: Work Name")
    print("     - Row 9: Work Order Number")
    print("     - Row 10: Agreement Number")
    print("     - Row 11: Work Order Amount")
    print("\n   Work Order Sheet:")
    print("     - Complete all item descriptions")
    print("     - Verify units, quantities, rates")
    print("     - Add any missing items")
    print("\n4. Save the file")
    print("\n5. Proceed to STEP 2: Generate Bill")
    print()
    
    return True


def main():
    """Main execution"""
    
    work_order_dir = "INPUT/work_order_samples/work_01_27022026"
    output_file = "OUTPUT/INPUT_work_01_27022026.xlsx"
    
    if len(sys.argv) > 1:
        work_order_dir = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    success = create_input_excel(work_order_dir, output_file)
    
    if success:
        print("\n✅ STEP 1 COMPLETE!")
        print("\nNext: Fill in the Excel file manually, then run STEP 2")
        print(f"      python process_first_bill.py {output_file}")
    
    return success


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
