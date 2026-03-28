#!/usr/bin/env python3
"""
Automated Excel Generator with Conflict Resolution
Creates professional Excel files with automatic error handling
"""

import os
import sys
import logging
import json
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime
import tempfile
import shutil

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..validation.data_validator import DataValidator, ValidationResult, ValidationLevel


@dataclass
class ExcelGenerationConfig:
    """Configuration for Excel generation"""
    output_dir: str = "OUTPUT"
    template_path: Optional[str] = None
    include_validation: bool = True
    include_formulas: bool = True
    apply_styling: bool = True
    backup_existing: bool = True
    timestamp_filename: bool = False


@dataclass
class GenerationResult:
    """Result of Excel generation"""
    success: bool
    file_path: str
    message: str
    backup_path: Optional[str] = None
    validation_report: Optional[Any] = None
    processing_time: float = 0.0


class ExcelConflictResolver:
    """Handles Excel file conflicts and locking issues"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def check_file_lock(self, file_path: str) -> bool:
        """Check if Excel file is locked"""
        try:
            # Try to open file in write mode
            with open(file_path, 'r+b') as f:
                pass
            return False
        except (PermissionError, OSError):
            return True
    
    def resolve_conflict(self, file_path: str, config: ExcelGenerationConfig) -> str:
        """Resolve file conflicts by creating backup or using timestamp"""
        if not os.path.exists(file_path):
            return file_path
        
        # Check if file is locked
        if self.check_file_lock(file_path):
            # Create backup and use timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{file_path}.backup_{timestamp}"
            
            try:
                shutil.copy2(file_path, backup_path)
                self.logger.info(f"Created backup: {backup_path}")
                
                # Create new file with timestamp
                new_path = f"{os.path.splitext(file_path)[0]}_{timestamp}{os.path.splitext(file_path)[1]}"
                return new_path
            except Exception as e:
                self.logger.error(f"Failed to create backup: {e}")
                raise
        
        # File exists but not locked
        if config.backup_existing:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{file_path}.backup_{timestamp}"
            shutil.copy2(file_path, backup_path)
            self.logger.info(f"Created backup: {backup_path}")
            return file_path
        
        return file_path


class ExcelStyler:
    """Applies professional styling to Excel files"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Define styles
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.data_font = Font(name='Calibri', size=11)
        self.error_font = Font(name='Calibri', size=11, color='FF0000')
        self.warning_font = Font(name='Calibri', size=11, color='FF9900')
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Alignment
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.currency_alignment = Alignment(horizontal='right', vertical='center')
    
    def style_worksheet(self, ws, data_type: str = 'data'):
        """Apply styling to worksheet"""
        # Style headers
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Style data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = self.data_font
                cell.border = self.thin_border
                
                # Currency alignment for numeric columns
                if cell.column_letter in ['E', 'F', 'G', 'H']:  # Rate, Quantity, Amount columns
                    cell.alignment = self.currency_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def apply_conditional_formatting(self, ws, validation_results: List[ValidationResult]):
        """Apply conditional formatting based on validation results"""
        # Red fill for errors
        error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Yellow fill for warnings  
        warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # Apply formatting based on validation results
        for result in validation_results:
            if result.level == ValidationLevel.ERROR:
                # Find cells with this value and apply error formatting
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value) == str(result.original_value):
                            cell.fill = error_fill
            
            elif result.level == ValidationLevel.WARNING:
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value) == str(result.original_value):
                            cell.fill = warning_fill


class AutomatedExcelGenerator:
    """Main automated Excel generator"""
    
    def __init__(self, config: Optional[ExcelGenerationConfig] = None):
        self.logger = logging.getLogger(__name__)
        self.config = config or ExcelGenerationConfig()
        self.conflict_resolver = ExcelConflictResolver()
        self.styler = ExcelStyler()
        self.validator = DataValidator()
        
        # Ensure output directory exists
        os.makedirs(self.config.output_dir, exist_ok=True)
    
    def generate_work_order_excel(self, items: List[Dict[str, Any]], 
                                filename: str = "INPUT_FINAL_WITH_QUANTITIES.xlsx",
                                qty_data: Optional[Dict[str, float]] = None) -> GenerationResult:
        """Generate work order Excel file with quantities"""
        start_time = datetime.now()
        
        try:
            # Validate data
            validation_report = self.validator.validate_dataset(items)
            
            # Apply corrections if requested
            if self.config.include_validation:
                items = self.validator.suggest_corrections(items)
            
            # Merge quantity data
            if qty_data:
                items = self._merge_quantities(items, qty_data)
            
            # Resolve file conflicts
            output_path = os.path.join(self.config.output_dir, filename)
            final_path = self.conflict_resolver.resolve_conflict(output_path, self.config)
            
            # Create Excel file
            wb = self._create_work_order_workbook(items, validation_report)
            
            # Save file
            wb.save(final_path)
            
            processing_time = (datetime.now() - start_time).total_seconds()
            
            return GenerationResult(
                success=True,
                file_path=final_path,
                message=f"Excel file generated successfully: {final_path}",
                validation_report=validation_report,
                processing_time=processing_time
            )
        
        except Exception as e:
            processing_time = (datetime.now() - start_time).total_seconds()
            self.logger.error(f"Error generating Excel file: {e}")
            
            return GenerationResult(
                success=False,
                file_path="",
                message=f"Error generating Excel file: {e}",
                processing_time=processing_time
            )
    
    def _create_work_order_workbook(self, items: List[Dict[str, Any]], 
                                  validation_report: Any) -> Workbook:
        """Create complete work order workbook"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_title_sheet(wb)
        self._create_work_order_sheet(wb, items)
        self._create_quantity_sheet(wb, items)
        self._create_bill_sheet(wb, items)
        
        if self.config.include_validation:
            self._create_validation_sheet(wb, validation_report)
        
        return wb
    
    def _create_title_sheet(self, wb: Workbook):
        """Create title sheet with project information"""
        ws = wb.create_sheet("TITLE", 0)
        
        # Project information
        title_data = [
            ["PWD CONTRACTOR BILL GENERATION SYSTEM", ""],
            ["", ""],
            ["Project Details", ""],
            ["Contractor Name:", "ENTER CONTRACTOR NAME"],
            ["Contract Number:", "ENTER CONTRACT NUMBER"],
            ["Work Order Number:", "ENTER WORK ORDER NUMBER"],
            ["Date:", datetime.now().strftime("%d-%m-%Y")],
            ["", ""],
            ["Generated By:", "Automated Bill Generator"],
            ["Generated On:", datetime.now().strftime("%d-%m-%Y %H:%M:%S")],
            ["Version:", "Production v2.0"]
        ]
        
        for row_data in title_data:
            ws.append(row_data)
        
        if self.config.apply_styling:
            self.styler.style_worksheet(ws, 'title')
    
    def _create_work_order_sheet(self, wb: Workbook, items: List[Dict[str, Any]]):
        """Create work order sheet with items"""
        ws = wb.create_sheet("Work Order")
        
        # Define columns
        columns = [
            "SN", "BSR Code", "Description", "Unit", "Quantity", 
            "Rate (INR)", "Amount (INR)", "Remarks"
        ]
        
        # Add header
        ws.append(columns)
        
        # Add items
        for i, item in enumerate(items, 1):
            quantity = item.get('quantity', 0)
            rate = item.get('rate', 0)
            amount = quantity * rate
            
            row = [
                i,
                item.get('bsr_code', ''),
                item.get('description', ''),
                item.get('unit', ''),
                quantity,
                rate,
                amount,
                item.get('remarks', '')
            ]
            ws.append(row)
        
        # Add totals
        last_row = len(items) + 2
        ws[f'F{last_row}'] = "TOTAL:"
        ws[f'G{last_row}'] = f"=SUM(G2:G{last_row-1})"
        
        if self.config.apply_styling:
            self.styler.style_worksheet(ws, 'data')
    
    def _create_quantity_sheet(self, wb: Workbook, items: List[Dict[str, Any]]):
        """Create quantity sheet"""
        ws = wb.create_sheet("Bill Quantity")
        
        # Define columns
        columns = ["SN", "BSR Code", "Description", "Unit", "Quantity", "Source"]
        
        # Add header
        ws.append(columns)
        
        # Add items with quantities
        for i, item in enumerate(items, 1):
            if item.get('quantity', 0) > 0:
                row = [
                    i,
                    item.get('bsr_code', ''),
                    item.get('description', ''),
                    item.get('unit', ''),
                    item.get('quantity', 0),
                    "OCR Extracted"
                ]
                ws.append(row)
        
        if self.config.apply_styling:
            self.styler.style_worksheet(ws, 'data')
    
    def _create_bill_sheet(self, wb: Workbook, items: List[Dict[str, Any]]):
        """Create final bill sheet"""
        ws = wb.create_sheet("Bill")
        
        # Bill calculations
        ws.append(["BILL CALCULATION", ""])
        ws.append(["", ""])
        
        # Add items for billing
        ws.append(["SN", "BSR Code", "Description", "Unit", "Quantity", "Rate", "Amount"])
        
        total_amount = 0
        for i, item in enumerate(items, 1):
            if item.get('quantity', 0) > 0:
                quantity = item.get('quantity', 0)
                rate = item.get('rate', 0)
                amount = quantity * rate
                total_amount += amount
                
                row = [i, item.get('bsr_code', ''), item.get('description', ''), 
                      item.get('unit', ''), quantity, rate, amount]
                ws.append(row)
        
        # Add summary
        ws.append(["", "", "", "", "", "SUBTOTAL:", total_amount])
        
        # Add common deductions (could be configurable)
        gst_rate = 0.18  # 18% GST
        gst_amount = total_amount * gst_rate
        
        ws.append(["", "", "", "", "", "GST (18%):", gst_amount])
        ws.append(["", "", "", "", "", "TOTAL BILL:", total_amount + gst_amount])
        
        if self.config.apply_styling:
            self.styler.style_worksheet(ws, 'bill')
    
    def _create_validation_sheet(self, wb: Workbook, validation_report: Any):
        """Create validation summary sheet"""
        ws = wb.create_sheet("Validation Report")
        
        # Validation summary
        ws.append(["VALIDATION REPORT", ""])
        ws.append(["", ""])
        
        # Summary statistics
        summary = validation_report.get_summary()
        ws.append(["Total Items:", summary['total_items']])
        ws.append(["Valid Items:", summary['valid_items']])
        ws.append(["Errors:", summary['error_count']])
        ws.append(["Warnings:", summary['warning_count']])
        ws.append(["Success Rate:", f"{summary['success_rate']:.1f}%"])
        ws.append(["Overall Score:", f"{summary['overall_score']:.1f}"])
        
        ws.append(["", ""])
        ws.append(["ERRORS:", ""])
        
        # List errors
        for error in validation_report.errors:
            ws.append([error.field, error.message])
        
        ws.append(["", ""])
        ws.append(["WARNINGS:", ""])
        
        # List warnings
        for warning in validation_report.warnings:
            ws.append([warning.field, warning.message])
        
        if self.config.apply_styling:
            self.styler.style_worksheet(ws, 'validation')
    
    def _merge_quantities(self, items: List[Dict[str, Any]], 
                         qty_data: Dict[str, float]) -> List[Dict[str, Any]]:
        """Merge quantity data from qty.txt with extracted items"""
        for item in items:
            bsr_code = item.get('bsr_code', '')
            
            # Direct match
            if bsr_code in qty_data:
                item['quantity'] = qty_data[bsr_code]
                continue
            
            # Partial match (e.g., 18.13 matches 18.13.6)
            for qty_bsr, qty in qty_data.items():
                if qty_bsr.startswith(bsr_code) or bsr_code.startswith(qty_bsr):
                    item['quantity'] = qty
                    break
        
        return items
    
    def generate_from_images(self, image_files: List[str], 
                           qty_file: Optional[str] = None) -> GenerationResult:
        """Generate Excel from image files (placeholder for OCR integration)"""
        # This would integrate with the OCR engine
        # For now, create a template structure
        items = []
        
        # Load quantity data
        qty_data = {}
        if qty_file and os.path.exists(qty_file):
            try:
                with open(qty_file, 'r') as f:
                    for line in f:
                        if ':' in line:
                            bsr, qty = line.strip().split(':', 1)
                            qty_data[bsr.strip()] = float(qty.strip())
            except Exception as e:
                self.logger.error(f"Error loading quantity file: {e}")
        
        return self.generate_work_order_excel(items, qty_data=qty_data)


def create_excel_generator(config: Optional[ExcelGenerationConfig] = None) -> AutomatedExcelGenerator:
    """Factory function to create Excel generator"""
    return AutomatedExcelGenerator(config)


if __name__ == "__main__":
    # Test the Excel generator
    logging.basicConfig(level=logging.INFO)
    
    generator = create_excel_generator()
    
    # Test items
    test_items = [
        {
            'bsr_code': '18.13',
            'description': 'Cement concrete 1:4:8 with PSC 40mm gauge',
            'unit': 'm3',
            'quantity': 10.5,
            'rate': 3500.00
        },
        {
            'bsr_code': '27.01',
            'description': 'First class brick masonry in cement mortar',
            'unit': 'm3',
            'quantity': 5.2,
            'rate': 4500.00
        }
    ]
    
    result = generator.generate_work_order_excel(test_items)
    print(f"Generation result: {result}")
