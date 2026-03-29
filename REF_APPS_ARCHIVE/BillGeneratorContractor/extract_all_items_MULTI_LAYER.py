#!/usr/bin/env python3
"""
Multi-Layer Extraction with Validation - Week 3 Complete
Gemini → Google Vision → EasyOCR + PWD Validation
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from modules.multi_layer_extractor import MultiLayerExtractor, ExtractionLayer
from modules.confidence_scorer import ConfidenceScorer
from modules.pwd_database import PWDDatabase
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Configuration
INPUT_FOLDER = Path("INPUT_WORK_ORDER_IMAGES_TEXT")
QTY_FILE = INPUT_FOLDER / "qty.txt"
OUTPUT_FILE = Path("OUTPUT/INPUT_MULTI_LAYER.xlsx")
API_KEY = "AIzaSyBMZYPgjcqXY-tpe6UhtBtrWhzfbU0-WVU"

def read_qty_file(qty_file_path):
    """Read quantities from qty.txt"""
    qty_data = {}
    with open(qty_file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                parts = line.split()
                if len(parts) >= 2:
                    item_no = parts[0]
                    quantity = float(parts[1])
                    qty_data[item_no] = quantity
    return qty_data

def sort_items_by_bsr(items):
    """Sort items by BSR code numerically"""
    def sort_key(item):
        code = item.get('code', '0')
        parts = code.split('.')
        try:
            return tuple(int(p) for p in parts)
        except:
            return (999, 999, 999)
    
    return sorted(items, key=sort_key)

def main():
    print("\n" + "="*80)
    print("MULTI-LAYER EXTRACTION - WEEK 3 COMPLETE")
    print("="*80)
    
    # Initialize components
    print("\nInitializing...")
    extractor = MultiLayerExtractor(gemini_api_key=API_KEY)
    db = PWDDatabase()
    scorer = ConfidenceScorer(db)
    
    # Show extractor status
    status = extractor.get_status()
    print(f"\nExtraction Layers:")
    for layer in status['layers']:
        available = "AVAILABLE" if layer['available'] else "NOT AVAILABLE"
        print(f"  - {layer['name']}: {available}")
    print(f"\nTotal available: {status['available_extractors']}/{status['total_extractors']}")
    
    # Get all images
    image_files = sorted(INPUT_FOLDER.glob("*.jpg")) + sorted(INPUT_FOLDER.glob("*.jpeg"))
    print(f"\nFound {len(image_files)} images to process")
    
    # Extract from all images with multi-layer fallback
    print("\n" + "="*80)
    print("EXTRACTING WITH MULTI-LAYER FALLBACK")
    print("="*80)
    
    all_items = []
    extraction_stats = {
        'gemini': 0,
        'google_vision': 0,
        'easyocr': 0,
        'failed': 0
    }
    
    for i, image_path in enumerate(image_files, 1):
        print(f"\n[{i}/{len(image_files)}] {image_path.name}")
        print("-" * 80)
        
        try:
            result = extractor.extract_with_fallback(str(image_path), min_confidence=0.7)
            
            if result.success and result.items:
                all_items.extend(result.items)
                extraction_stats[result.layer.value] += 1
                print(f"\n  SUCCESS: Extracted {len(result.items)} items using {result.layer.value}")
            else:
                extraction_stats['failed'] += 1
                print(f"\n  FAILED: No items extracted")
        
        except Exception as e:
            extraction_stats['failed'] += 1
            print(f"\n  ERROR: {e}")
    
    # Remove duplicates by BSR code
    unique_items = {}
    for item in all_items:
        code = item.get('code', '')
        if code and code not in unique_items:
            unique_items[code] = item
    
    items_list = list(unique_items.values())
    print(f"\n" + "="*80)
    print(f"EXTRACTION COMPLETE")
    print("="*80)
    print(f"Total items extracted: {len(all_items)}")
    print(f"Unique items: {len(items_list)}")
    print(f"\nExtraction by layer:")
    for layer, count in extraction_stats.items():
        if count > 0:
            percentage = (count / len(image_files)) * 100
            print(f"  {layer.upper():15s}: {count:2d} images ({percentage:5.1f}%)")
    
    # Sort by BSR code
    items_list = sort_items_by_bsr(items_list)
    
    # Validate all items
    print("\n" + "="*80)
    print("VALIDATING ITEMS")
    print("="*80)
    
    scores = scorer.score_items(items_list)
    report = scorer.generate_report(items_list)
    
    print(f"\nValidation Summary:")
    print(f"  Total Items: {report['total_items']}")
    print(f"  Average Confidence: {report['average_confidence']:.2f}")
    print(f"\nConfidence Distribution:")
    for level, data in report['confidence_levels'].items():
        if data['count'] > 0:
            print(f"  {level.upper():12s}: {data['count']:2d} items ({data['percentage']:5.1f}%)")
    
    # Read quantities
    print("\n" + "="*80)
    print("READING QUANTITIES")
    print("="*80)
    
    qty_data = read_qty_file(QTY_FILE)
    print(f"Loaded {len(qty_data)} quantities from {QTY_FILE.name}")
    
    # Create Excel
    print("\n" + "="*80)
    print("CREATING EXCEL")
    print("="*80)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["MULTI-LAYER EXTRACTION REPORT"])
    ws_summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])
    ws_summary.append(["EXTRACTION STATISTICS"])
    ws_summary.append(["Total Images:", len(image_files)])
    ws_summary.append(["Total Items:", len(all_items)])
    ws_summary.append(["Unique Items:", len(items_list)])
    ws_summary.append([])
    ws_summary.append(["Extraction by Layer:"])
    for layer, count in extraction_stats.items():
        if count > 0:
            percentage = (count / len(image_files)) * 100
            ws_summary.append([f"  {layer.upper()}", count, f"{percentage:.1f}%"])
    ws_summary.append([])
    ws_summary.append(["VALIDATION STATISTICS"])
    ws_summary.append(["Average Confidence:", f"{report['average_confidence']:.2f}"])
    ws_summary.append([])
    ws_summary.append(["Confidence Distribution:"])
    for level, data in report['confidence_levels'].items():
        if data['count'] > 0:
            ws_summary.append([f"  {level.upper()}", data['count'], f"{data['percentage']:.1f}%"])
    
    # Sheet 2: Work Order
    ws_work = wb.create_sheet("Work Order")
    headers = ["Item", "Description", "Unit", "Quantity", "Rate", "Amount", "BSR", "Confidence", "Status"]
    ws_work.append(headers)
    
    for cell in ws_work[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for i, (item, score) in enumerate(zip(items_list, scores), 1):
        code = item.get('code', '')
        desc = item.get('description', '')
        unit = item.get('unit', '')
        qty = item.get('quantity', 1)
        rate = item.get('rate', 0)
        amount = qty * rate
        
        if score.overall >= 0.95:
            status = "AUTO_ACCEPT"
            fill_color = "C6EFCE"
        elif score.overall >= 0.85:
            status = "QUICK_REVIEW"
            fill_color = "FFEB9C"
        elif score.overall >= 0.70:
            status = "REVIEW"
            fill_color = "FFC7CE"
        else:
            status = "DETAILED_REVIEW"
            fill_color = "FF6B6B"
        
        row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}", status]
        ws_work.append(row)
        
        for cell in ws_work[i+1]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Sheet 3: Bill Quantity
    ws_bill = wb.create_sheet("Bill Quantity")
    ws_bill.append(headers)
    
    for cell in ws_bill[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for i, (item, score) in enumerate(zip(items_list, scores), 1):
        code = item.get('code', '')
        desc = item.get('description', '')
        unit = item.get('unit', '')
        rate = item.get('rate', 0)
        qty = qty_data.get(code, 0)
        amount = qty * rate
        
        if score.overall >= 0.95:
            status = "AUTO_ACCEPT"
            fill_color = "C6EFCE"
        elif score.overall >= 0.85:
            status = "QUICK_REVIEW"
            fill_color = "FFEB9C"
        elif score.overall >= 0.70:
            status = "REVIEW"
            fill_color = "FFC7CE"
        else:
            status = "DETAILED_REVIEW"
            fill_color = "FF6B6B"
        
        row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}", status]
        ws_bill.append(row)
        
        for cell in ws_bill[i+1]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Save
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)
    
    print(f"\nExcel file created: {OUTPUT_FILE}")
    
    # Final statistics
    print("\n" + "="*80)
    print("WEEK 3 COMPLETE - MULTI-LAYER EXTRACTION")
    print("="*80)
    print(f"\nReliability Metrics:")
    print(f"  Extraction Success: {((len(image_files) - extraction_stats['failed']) / len(image_files) * 100):.1f}%")
    print(f"  Validation Confidence: {report['average_confidence']*100:.1f}%")
    print(f"  Auto-Accept Rate: {report['recommended_actions']['auto_accept']['percentage']:.1f}%")
    print(f"\nLayer Distribution:")
    for layer, count in extraction_stats.items():
        if count > 0:
            percentage = (count / len(image_files)) * 100
            print(f"  {layer.upper():15s}: {percentage:5.1f}%")
    print("="*80 + "\n")

if __name__ == '__main__':
    main()
