#!/usr/bin/env python3
"""
Create properly formatted Work Order Excel file matching TEST_INPUT_FILES structure
Based on work order images in INPUT/work_order_samples/work_01_27022026
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_formatted_work_order(output_file: str = "OUTPUT/work_order_from_images.xlsx"):
    """
    Create Excel file with exact structure matching TEST_INPUT_FILES
    """
    
    print(f"\n{'='*80}")
    print(f"Creating Formatted Work Order Excel File")
    print(f"{'='*80}\n")
    print("This file will match the structure of TEST_INPUT_FILES")
    print("You need to fill in the data from your work order images\n")
    
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ========================================================================
    # SHEET 1: Title (Matching FirstFINALnoExtra.xlsx structure)
    # ========================================================================
    print("Creating Title sheet...")
    ws_title = wb.create_sheet("Title")
    
    # Title sheet data - Column A has labels, Column B has values
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FROM GOVT. TREASURY", ""],
        ["Bill Number", ""],  # Fill from images
        ["Running or Final", "Final"],  # Or "Running" or "First"
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", ""],  # Fill from images
        ["Name of Work ;-", ""],  # Fill from images
        ["Serial No. of this bill :", ""],  # e.g., "First & Final Bill"
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", ""],  # Fill from images
        ["Agreement No.", ""],  # Fill from images
        ["WORK ORDER AMOUNT RS.", ""],  # Fill from images
        ["Date of written order to commence work :", ""],  # Format: YYYY-MM-DD
        ["St. date of Start :", ""],  # Format: YYYY-MM-DD
        ["St. date of completion :", ""],  # Format: YYYY-MM-DD
        ["Date of actual completion of work :", ""],  # Format: YYYY-MM-DD
        ["Date of measurement :", ""],  # Format: YYYY-MM-DD
        ["TENDER PREMIUM %", ""],  # e.g., 11.22
        ["Above / Below", "Above"],  # Or "Below"
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        ws_title.cell(row=row_idx, column=1, value=label)
        ws_title.cell(row=row_idx, column=2, value=value)
    
    # Format Title sheet
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # ========================================================================
    # SHEET 2: Work Order (Matching FirstFINALnoExtra.xlsx structure)
    # ========================================================================
    print("Creating Work Order sheet...")
    ws_work = wb.create_sheet("Work Order")
    
    # Headers
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_work.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Sample rows - YOU NEED TO FILL THESE FROM YOUR IMAGES
    sample_data = [
        # [Item, Description, Unit, Quantity, Rate, Amount, BSR]
        [1.0, "ITEM DESCRIPTION FROM IMAGE 1", "sqm", "", "", "", ""],
        ["", "  Sub-item or continuation", "sqm", "", "", "", ""],
        [2.0, "ITEM DESCRIPTION FROM IMAGE 2", "cum", "", "", "", ""],
        [3.0, "ITEM DESCRIPTION FROM IMAGE 3", "kg", "", "", "", ""],
        [4.0, "ITEM DESCRIPTION FROM IMAGE 4", "nos", "", "", "", ""],
        [5.0, "ITEM DESCRIPTION FROM IMAGE 5", "rmt", "", "", "", ""],
        [6.0, "ITEM DESCRIPTION FROM IMAGE 6", "mt", "", "", "", ""],
        [7.0, "ITEM DESCRIPTION FROM IMAGE 7", "ltr", "", "", "", ""],
        [8.0, "ITEM DESCRIPTION FROM IMAGE 8", "each", "", "", "", ""],
        [9.0, "ITEM DESCRIPTION FROM IMAGE 9", "set", "", "", "", ""],
        [10.0, "ITEM DESCRIPTION FROM IMAGE 10", "sqm", "", "", "", ""],
        # Add more rows as needed from your images
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_work.cell(row=row_idx, column=col_idx, value=value)
    
    # Format Work Order sheet
    ws_work.column_dimensions['A'].width = 8
    ws_work.column_dimensions['B'].width = 60
    ws_work.column_dimensions['C'].width = 10
    ws_work.column_dimensions['D'].width = 12
    ws_work.column_dimensions['E'].width = 12
    ws_work.column_dimensions['F'].width = 15
    ws_work.column_dimensions['G'].width = 10
    
    # ========================================================================
    # SHEET 3: Bill Quantity (Copy of Work Order for editing)
    # ========================================================================
    print("Creating Bill Quantity sheet...")
    ws_bill = wb.create_sheet("Bill Quantity")
    
    # Copy headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bill.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Copy sample data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_bill.cell(row=row_idx, column=col_idx, value=value)
    
    # Format Bill Quantity sheet
    ws_bill.column_dimensions['A'].width = 8
    ws_bill.column_dimensions['B'].width = 60
    ws_bill.column_dimensions['C'].width = 10
    ws_bill.column_dimensions['D'].width = 12
    ws_bill.column_dimensions['E'].width = 12
    ws_bill.column_dimensions['F'].width = 15
    ws_bill.column_dimensions['G'].width = 10
    
    # ========================================================================
    # SHEET 4: Extra Items
    # ========================================================================
    print("Creating Extra Items sheet...")
    ws_extra = wb.create_sheet("Extra Items")
    
    # Headers for Extra Items
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        cell = ws_extra.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Sample extra items (leave empty or add if needed)
    extra_sample = [
        ["E1", "Extra item description (if any)", "", "", "", "", "", ""],
        ["E2", "Extra item description (if any)", "", "", "", "", "", ""],
        ["E3", "Extra item description (if any)", "", "", "", "", "", ""],
    ]
    
    for row_idx, row_data in enumerate(extra_sample, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_extra.cell(row=row_idx, column=col_idx, value=value)
    
    # Format Extra Items sheet
    ws_extra.column_dimensions['A'].width = 8
    ws_extra.column_dimensions['B'].width = 60
    ws_extra.column_dimensions['C'].width = 10
    ws_extra.column_dimensions['D'].width = 12
    ws_extra.column_dimensions['E'].width = 12
    ws_extra.column_dimensions['F'].width = 15
    ws_extra.column_dimensions['G'].width = 12
    ws_extra.column_dimensions['H'].width = 10
    
    # Save workbook
    wb.save(output_path)
    
    print(f"\n✅ Excel file created: {output_path.absolute()}\n")
    
    # Create detailed instructions
    instructions_file = output_path.parent / "FILLING_INSTRUCTIONS.txt"
    with open(instructions_file, 'w', encoding='utf-8') as f:
        f.write("WORK ORDER DATA ENTRY INSTRUCTIONS\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Excel File: {output_path.name}\n")
        f.write(f"Work Order Images: INPUT/work_order_samples/work_01_27022026/\n\n")
        
        f.write("STEP-BY-STEP GUIDE\n")
        f.write("=" * 80 + "\n\n")
        
        f.write("STEP 1: Open Both Files\n")
        f.write("-" * 80 + "\n")
        f.write(f"1. Open Excel file: {output_path.name}\n")
        f.write("2. Open image folder: INPUT/work_order_samples/work_01_27022026/\n")
        f.write("3. Arrange windows side-by-side\n\n")
        
        f.write("STEP 2: Fill Title Sheet (Column B)\n")
        f.write("-" * 80 + "\n")
        f.write("Look at Image 1 (WhatsApp Image ... 1.13.49 PM.jpeg) for:\n\n")
        f.write("Row 2:  Bill Number - Enter bill number\n")
        f.write("Row 3:  Running or Final - Keep 'Final' or change to 'Running'/'First'\n")
        f.write("Row 5:  Contractor Name - Enter contractor name from image\n")
        f.write("Row 6:  Name of Work - Enter work description from image\n")
        f.write("Row 7:  Serial No. - e.g., 'First & Final Bill'\n")
        f.write("Row 9:  Work Order Reference - Enter work order number and date\n")
        f.write("Row 10: Agreement No. - Enter agreement number\n")
        f.write("Row 11: Work Order Amount - Enter total amount\n")
        f.write("Row 12-16: Dates - Enter in YYYY-MM-DD format (e.g., 2025-01-09)\n")
        f.write("Row 17: Tender Premium % - Enter percentage (e.g., 11.22)\n")
        f.write("Row 18: Above/Below - Keep 'Above' or change to 'Below'\n\n")
        
        f.write("STEP 3: Fill Work Order Sheet\n")
        f.write("-" * 80 + "\n")
        f.write("Look at ALL images for work items:\n\n")
        f.write("For EACH item in the work order:\n")
        f.write("  Column A (Item): Item number (1.0, 2.0, 3.0, ...)\n")
        f.write("  Column B (Description): Full item description from image\n")
        f.write("  Column C (Unit): Unit of measurement (sqm, cum, kg, nos, rmt, etc.)\n")
        f.write("  Column D (Quantity): Quantity from work order\n")
        f.write("  Column E (Rate): Rate per unit\n")
        f.write("  Column F (Amount): Quantity × Rate\n")
        f.write("  Column G (BSR): BSR code if available\n\n")
        
        f.write("IMPORTANT NOTES:\n")
        f.write("  - If item has sub-items, leave Column A empty for sub-items\n")
        f.write("  - Keep descriptions exactly as in work order\n")
        f.write("  - Add more rows if needed (right-click → Insert)\n")
        f.write("  - Delete sample rows that you don't need\n\n")
        
        f.write("STEP 4: Fill Bill Quantity Sheet\n")
        f.write("-" * 80 + "\n")
        f.write("1. Copy all items from Work Order sheet\n")
        f.write("2. Update Column D (Quantity) with actual work done\n")
        f.write("3. Keep rates same as Work Order\n")
        f.write("4. Recalculate Column F (Amount) = Quantity × Rate\n\n")
        
        f.write("STEP 5: Fill Extra Items Sheet (If Applicable)\n")
        f.write("-" * 80 + "\n")
        f.write("Only if you have extra items NOT in original work order:\n")
        f.write("  Column A: E1, E2, E3, ... (Extra item numbers)\n")
        f.write("  Column B: Description of extra item\n")
        f.write("  Column C-F: Unit, Quantity, Rate, Amount\n")
        f.write("  Column G: Deviation percentage\n")
        f.write("  Column H: BSR code if available\n\n")
        
        f.write("STEP 6: Save and Process\n")
        f.write("-" * 80 + "\n")
        f.write("1. Save the Excel file (Ctrl+S)\n")
        f.write("2. Close Excel\n")
        f.write("3. Run command:\n")
        f.write(f"   python process_first_bill.py {output_path}\n\n")
        
        f.write("COMMON UNITS\n")
        f.write("-" * 80 + "\n")
        f.write("sqm  - Square meter (area)\n")
        f.write("cum  - Cubic meter (volume)\n")
        f.write("kg   - Kilogram (weight)\n")
        f.write("nos  - Numbers (count)\n")
        f.write("rmt  - Running meter (length)\n")
        f.write("mt   - Metric ton\n")
        f.write("ltr  - Liter (volume)\n")
        f.write("each - Each item\n")
        f.write("set  - Set of items\n\n")
        
        f.write("TIPS FOR ACCURACY\n")
        f.write("-" * 80 + "\n")
        f.write("✓ Double-check all numbers\n")
        f.write("✓ Verify calculations (Amount = Quantity × Rate)\n")
        f.write("✓ Keep item numbers sequential\n")
        f.write("✓ Copy descriptions exactly from images\n")
        f.write("✓ Use consistent units throughout\n")
        f.write("✓ Save frequently while working\n\n")
        
        f.write("NEED HELP?\n")
        f.write("-" * 80 + "\n")
        f.write("Refer to:\n")
        f.write("  - VIEW_WORK_ORDER_IMAGES.md - Image viewing guide\n")
        f.write("  - COMPLETE_WORKFLOW_GUIDE.md - Complete workflow\n")
        f.write("  - TEST_INPUT_FILES/FirstFINALnoExtra.xlsx - Example file\n")
    
    print(f"✅ Instructions created: {instructions_file.absolute()}\n")
    
    print(f"{'='*80}")
    print("✅ File Creation Complete!")
    print(f"{'='*80}\n")
    print("This Excel file matches the exact structure of TEST_INPUT_FILES")
    print("\nNext steps:")
    print("1. Open the Excel file")
    print("2. Open the work order images")
    print("3. Fill in the data from images")
    print("4. Save the file")
    print(f"5. Run: python process_first_bill.py {output_path}")
    print()
    
    return output_path


if __name__ == '__main__':
    import sys
    
    output_file = "OUTPUT/work_order_from_images.xlsx"
    if len(sys.argv) > 1:
        output_file = sys.argv[1]
    
    result = create_formatted_work_order(output_file)
    
    # Open the file and images
    print("Opening files for you...")
    import subprocess
    subprocess.run(f'start {result}', shell=True)
    subprocess.run('start INPUT\\work_order_samples\\work_01_27022026\\', shell=True)
    print("\n✅ Excel file and images folder opened!")
    print("You can now start filling in the data.\n")
