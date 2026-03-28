#!/usr/bin/env python3
"""
Helper script to update Title sheet with actual work order details
Since OCR is not available, you need to manually enter data from images
"""
import openpyxl
from pathlib import Path

def update_title_sheet():
    """Update Title sheet with actual work order data"""
    
    excel_file = Path("OUTPUT/INPUT_work_01_27022026_PRODUCTION.xlsx")
    
    if not excel_file.exists():
        print("❌ Excel file not found. Run create_excel_production.py first.")
        return
    
    print("="*80)
    print("UPDATE TITLE SHEET - Manual Data Entry")
    print("="*80)
    print()
    print("Please look at the work order images and enter the following details:")
    print()
    
    # Get user input
    contractor_name = input("Contractor Name (M/s. ...): ").strip()
    if not contractor_name.startswith("M/s."):
        contractor_name = f"M/s. {contractor_name}"
    
    work_name = input("Name of Work: ").strip()
    work_order_no = input("Work Order Number: ").strip()
    agreement_no = input("Agreement Number: ").strip()
    work_order_amount = input("Work Order Amount (Rs.): ").strip()
    
    # Optional dates
    print()
    print("Dates (press Enter to keep defaults):")
    date_commence = input("Date of written order to commence work (YYYY-MM-DD) [2026-02-27]: ").strip() or "2026-02-27"
    date_start = input("Start date (YYYY-MM-DD) [2026-03-01]: ").strip() or "2026-03-01"
    date_completion_scheduled = input("Scheduled completion date (YYYY-MM-DD) [2026-06-30]: ").strip() or "2026-06-30"
    date_completion_actual = input("Actual completion date (YYYY-MM-DD) [2026-06-30]: ").strip() or "2026-06-30"
    
    # Optional premium
    print()
    tender_premium = input("Tender Premium % [11.22]: ").strip() or "11.22"
    premium_type = input("Above/Below [Above]: ").strip() or "Above"
    
    # Update Excel
    print()
    print("Updating Excel file...")
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Title']
    
    # Update values (row numbers based on create_excel_production.py)
    ws.cell(5, 2, contractor_name)  # Contractor name
    ws.cell(6, 2, work_name)  # Work name
    ws.cell(9, 2, work_order_no)  # Work order number
    ws.cell(10, 2, agreement_no)  # Agreement number
    ws.cell(11, 2, work_order_amount)  # Work order amount
    ws.cell(12, 2, date_commence)  # Date of written order
    ws.cell(13, 2, date_start)  # Start date
    ws.cell(14, 2, date_completion_scheduled)  # Scheduled completion
    ws.cell(15, 2, date_completion_actual)  # Actual completion
    ws.cell(17, 2, tender_premium)  # Tender premium
    ws.cell(18, 2, premium_type)  # Above/Below
    
    wb.save(excel_file)
    
    print("✅ Title sheet updated successfully!")
    print()
    print("="*80)
    print("UPDATED DETAILS:")
    print("="*80)
    print(f"Contractor: {contractor_name}")
    print(f"Work Name: {work_name}")
    print(f"Work Order No: {work_order_no}")
    print(f"Agreement No: {agreement_no}")
    print(f"Work Order Amount: Rs. {work_order_amount}")
    print(f"Tender Premium: {tender_premium}% {premium_type}")
    print()
    print(f"File saved: {excel_file.absolute()}")
    print()
    print("Next step: Generate bill documents")
    print(f"  python process_first_bill.py {excel_file}")
    print()

if __name__ == '__main__':
    try:
        update_title_sheet()
    except KeyboardInterrupt:
        print("\n\n⚠️  Cancelled by user")
    except Exception as e:
        print(f"\n❌ Error: {e}")
