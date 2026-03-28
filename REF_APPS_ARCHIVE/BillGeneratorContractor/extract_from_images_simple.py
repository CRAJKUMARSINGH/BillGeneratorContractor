#!/usr/bin/env python3
"""
SIMPLE IMAGE EXTRACTION - Process ALL 5 images and extract ALL text
Then create INPUT Excel matching TEST_INPUT format
"""
import sys
from pathlib import Path
import easyocr
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

def extract_text_from_all_images(work_order_dir):
    """Extract text from all images using EasyOCR"""
    
    work_order_path = Path(work_order_dir)
    
    # Find all images
    image_files = []
    for ext in ['*.jpg', '*.jpeg', '*.png', '*.JPG', '*.JPEG', '*.PNG']:
        image_files.extend(sorted(work_order_path.glob(ext)))
    
    if not image_files:
        print("❌ No images found!")
        return []
    
    print(f"\n📸 Found {len(image_files)} images")
    for img in image_files:
        print(f"   - {img.name}")
    
    # Initialize EasyOCR
    print("\n🔧 Initializing EasyOCR...")
    reader = easyocr.Reader(['en', 'hi'], gpu=False)
    print("✅ EasyOCR ready\n")
    
    # Extract text from each image
    all_text = []
    
    for idx, img_file in enumerate(image_files, 1):
        print(f"📄 Processing image {idx}/{len(image_files)}: {img_file.name}")
        
        try:
            # Perform OCR
            result = reader.readtext(str(img_file))
            
            # Extract text
            image_text = []
            for detection in result:
                text = detection[1]
                confidence = detection[2]
                if confidence > 0.3:  # Low threshold to capture everything
                    image_text.append(text)
            
            print(f"   ✅ Extracted {len(image_text)} text blocks")
            
            # Store with image reference
            all_text.append({
                'image': img_file.name,
                'text_blocks': image_text,
                'full_text': ' '.join(image_text)
            })
            
        except Exception as e:
            print(f"   ❌ Error: {e}")
    
    return all_text

def create_input_excel_from_text(all_text, qty_file, output_file):
    """Create INPUT Excel from extracted text + qty.txt"""
    
    print(f"\n{'='*80}")
    print("📊 CREATING INPUT EXCEL FROM EXTRACTED TEXT")
    print(f"{'='*80}\n")
    
    # Read qty.txt
    qty_data = {}
    if Path(qty_file).exists():
        with open(qty_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split()
                    if len(parts) >= 2:
                        code = parts[0]
                        qty = float(parts[1])
                        qty_data[code] = qty
        print(f"✅ Loaded {len(qty_data)} items from qty.txt\n")
    
    # Combine all text
    combined_text = '\n\n'.join([f"=== {item['image']} ===\n{item['full_text']}" for item in all_text])
    
    # Save extracted text for reference
    text_output = Path(output_file).parent / "EXTRACTED_TEXT_ALL_IMAGES.txt"
    with open(text_output, 'w', encoding='utf-8') as f:
        f.write(combined_text)
    print(f"✅ Saved extracted text to: {text_output}\n")
    
    # Create Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # === SHEET 1: Title ===
    ws_title = wb.create_sheet("Title")
    title_data = [
        ["FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED", ""],
        ["Bill Number", "First"],
        ["Running or Final", "Final"],
        ["Cash Book Voucher No. and Date", ""],
        ["Name of Contractor or supplier :", "[CHECK EXTRACTED_TEXT_ALL_IMAGES.txt]"],
        ["Name of Work ;-", "[CHECK EXTRACTED_TEXT_ALL_IMAGES.txt]"],
        ["Serial No. of this bill :", "First & Final Bill"],
        ["No. and date of the last bill-", "Not Applicable"],
        ["Reference to work order or Agreement :", "[CHECK EXTRACTED_TEXT_ALL_IMAGES.txt]"],
        ["Agreement No.", "[CHECK EXTRACTED_TEXT_ALL_IMAGES.txt]"],
        ["WORK ORDER AMOUNT RS.", "[CHECK EXTRACTED_TEXT_ALL_IMAGES.txt]"],
        ["Date of written order to commence work :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of Start :", datetime.now().strftime("%Y-%m-%d")],
        ["St. date of completion :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of actual completion of work :", datetime.now().strftime("%Y-%m-%d")],
        ["Date of measurement :", datetime.now().strftime("%Y-%m-%d")],
        ["TENDER PREMIUM %", "11.22"],
        ["Above / Below", "Above"],
        ["Amount Paid Vide Last Bill", "0"]
    ]
    
    for row_idx, (label, value) in enumerate(title_data, start=1):
        cell_label = ws_title.cell(row=row_idx, column=1, value=label)
        cell_value = ws_title.cell(row=row_idx, column=2, value=value)
        cell_label.font = Font(bold=True)
    
    ws_title.column_dimensions['A'].width = 50
    ws_title.column_dimensions['B'].width = 40
    
    # === SHEET 2: Work Order ===
    ws_work = wb.create_sheet("Work Order")
    headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'BSR']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_work.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add placeholder row
    ws_work.cell(row=2, column=1, value="[MANUALLY FILL FROM EXTRACTED_TEXT_ALL_IMAGES.txt]")
    ws_work.cell(row=2, column=2, value="[Item descriptions from images]")
    ws_work.cell(row=2, column=3, value="[Unit]")
    ws_work.cell(row=2, column=4, value="[Qty]")
    ws_work.cell(row=2, column=5, value="[Rate]")
    ws_work.cell(row=2, column=6, value="[Amount]")
    ws_work.cell(row=2, column=7, value="[BSR]")
    
    ws_work.column_dimensions['B'].width = 80
    
    # === SHEET 3: Bill Quantity ===
    ws_bill = wb.create_sheet("Bill Quantity")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_bill.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add items from qty.txt
    row_idx = 2
    for code, qty in sorted(qty_data.items()):
        ws_bill.cell(row=row_idx, column=1, value=code)
        ws_bill.cell(row=row_idx, column=2, value=f"[Find description for {code} in EXTRACTED_TEXT_ALL_IMAGES.txt]")
        ws_bill.cell(row=row_idx, column=3, value="[Unit]")
        ws_bill.cell(row=row_idx, column=4, value=qty)
        ws_bill.cell(row=row_idx, column=5, value="[Rate]")
        ws_bill.cell(row=row_idx, column=6, value="[Qty × Rate]")
        ws_bill.cell(row=row_idx, column=7, value=code)
        row_idx += 1
    
    ws_bill.column_dimensions['B'].width = 80
    
    # === SHEET 4: Extra Items ===
    ws_extra = wb.create_sheet("Extra Items")
    extra_headers = ['Item', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount', 'Deviation %', 'BSR']
    for col_idx, header in enumerate(extra_headers, start=1):
        cell = ws_extra.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    ws_extra.column_dimensions['B'].width = 80
    
    # Save Excel
    wb.save(output_file)
    
    print(f"{'='*80}")
    print("✅ INPUT EXCEL CREATED")
    print(f"{'='*80}\n")
    print(f"📁 Excel: {output_file}")
    print(f"📄 Text: {text_output}")
    print(f"\n📋 Next Steps:")
    print(f"1. Open {text_output}")
    print(f"2. Find item descriptions, rates, units from the extracted text")
    print(f"3. Fill in the Excel file manually")
    print(f"4. Or use the text to improve OCR extraction")
    
    return True

def main():
    work_order_dir = "INPUT/work_order_samples/work_01_27022026"
    output_file = "OUTPUT/INPUT_WITH_EXTRACTED_TEXT.xlsx"
    
    if len(sys.argv) > 1:
        work_order_dir = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    # Extract text from all images
    all_text = extract_text_from_all_images(work_order_dir)
    
    if not all_text:
        print("❌ No text extracted from images")
        return 1
    
    # Create Excel
    qty_file = Path(work_order_dir) / "qty.txt"
    success = create_input_excel_from_text(all_text, qty_file, output_file)
    
    return 0 if success else 1

if __name__ == '__main__':
    sys.exit(main())
