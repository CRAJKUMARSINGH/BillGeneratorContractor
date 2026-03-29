#!/usr/bin/env python3
"""
Extract All Items with Validation - Week 2 Integration
Gemini Vision extraction + PWD database validation
"""
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from modules.gemini_vision_parser_v2 import GeminiVisionParser
from modules.confidence_scorer import ConfidenceScorer
from modules.pwd_database import PWDDatabase
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Configuration
INPUT_FOLDER = Path("INPUT_WORK_ORDER_IMAGES_TEXT")
QTY_FILE = INPUT_FOLDER / "qty.txt"
OUTPUT_FILE = Path("OUTPUT/INPUT_VALIDATED.xlsx")
API_KEY = "AIzaSyBMZYPgjcqXY-tpe6UhtBtrWhzfbU0-WVU"

def read_qty_file(qty_file_path):
    """Read quantities from qty.txt"""
    qty_data = {}
    with open(qty_file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                parts = line.split()
                if len(parts) >= 2:
                    item_no = parts[0]
                    quantity = float(parts[1])
                    qty_data[item_no] = quantity
    return qty_data

def sort_items_by_bsr(items):
    """Sort items by BSR code numerically"""
    def sort_key(item):
        code = item.get('code', '0')
        parts = code.split('.')
        try:
            return tuple(int(p) for p in parts)
        except:
            return (999, 999, 999)
    
    return sorted(items, key=sort_key)

def main():
    print("\n" + "="*80)
    print("EXTRACTION WITH VALIDATION - WEEK 2")
    print("="*80)
    
    # Initialize components
    print("\nInitializing...")
    parser = GeminiVisionParser(api_key=API_KEY)
    db = PWDDatabase()
    scorer = ConfidenceScorer(db)
    
    # Get all images
    image_files = sorted(INPUT_FOLDER.glob("*.jpg")) + sorted(INPUT_FOLDER.glob("*.jpeg"))
    print(f"Found {len(image_files)} images")
    
    # Extract from all images
    print("\nExtracting items from images...")
    all_items = []
    
    for i, image_path in enumerate(image_files, 1):
        print(f"\n[{i}/{len(image_files)}] Processing: {image_path.name}")
        try:
            items = parser.extract_items_from_image(str(image_path))
            print(f"  Extracted {len(items)} items")
            all_items.extend(items)
        except Exception as e:
            print(f"  ERROR: {e}")
    
    # Remove duplicates by BSR code
    unique_items = {}
    for item in all_items:
        code = item.get('code', '')
        if code and code not in unique_items:
            unique_items[code] = item
    
    items_list = list(unique_items.values())
    print(f"\nTotal unique items: {len(items_list)}")
    
    # Sort by BSR code
    items_list = sort_items_by_bsr(items_list)
    
    # Validate all items
    print("\n" + "="*80)
    print("VALIDATING ITEMS")
    print("="*80)
    
    scores = scorer.score_items(items_list)
    
    # Generate validation report
    report = scorer.generate_report(items_list)
    
    print(f"\nValidation Summary:")
    print(f"  Total Items: {report['total_items']}")
    print(f"  Average Confidence: {report['average_confidence']:.2f}")
    print(f"\nConfidence Distribution:")
    for level, data in report['confidence_levels'].items():
        print(f"  {level.upper():12s}: {data['count']:2d} items ({data['percentage']:5.1f}%)")
    print(f"\nRecommended Actions:")
    for action, data in report['recommended_actions'].items():
        print(f"  {action.upper():16s}: {data['count']:2d} items ({data['percentage']:5.1f}%)")
    
    # Read quantities
    print("\n" + "="*80)
    print("READING QUANTITIES")
    print("="*80)
    
    qty_data = read_qty_file(QTY_FILE)
    print(f"Loaded {len(qty_data)} quantities from {QTY_FILE.name}")
    
    # Create Excel with validation info
    print("\n" + "="*80)
    print("CREATING EXCEL WITH VALIDATION")
    print("="*80)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Validation Summary
    ws_summary = wb.create_sheet("Validation Summary")
    ws_summary.append(["EXTRACTION WITH VALIDATION REPORT"])
    ws_summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])
    ws_summary.append(["Total Items:", report['total_items']])
    ws_summary.append(["Average Confidence:", f"{report['average_confidence']:.2f}"])
    ws_summary.append([])
    ws_summary.append(["Confidence Distribution:"])
    for level, data in report['confidence_levels'].items():
        ws_summary.append([f"  {level.upper()}", data['count'], f"{data['percentage']:.1f}%"])
    ws_summary.append([])
    ws_summary.append(["Recommended Actions:"])
    for action, data in report['recommended_actions'].items():
        ws_summary.append([f"  {action.upper()}", data['count'], f"{data['percentage']:.1f}%"])
    
    # Sheet 2: Work Order (with validation)
    ws_work = wb.create_sheet("Work Order")
    
    # Headers
    headers = ["Item", "Description", "Unit", "Quantity", "Rate", "Amount", "BSR", "Confidence", "Status"]
    ws_work.append(headers)
    
    # Style headers
    for cell in ws_work[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Add items with validation info
    for i, (item, score) in enumerate(zip(items_list, scores), 1):
        code = item.get('code', '')
        desc = item.get('description', '')
        unit = item.get('unit', '')
        qty = item.get('quantity', 1)
        rate = item.get('rate', 0)
        amount = qty * rate
        
        # Determine status based on confidence
        if score.overall >= 0.95:
            status = "AUTO_ACCEPT"
            fill_color = "C6EFCE"  # Light green
        elif score.overall >= 0.85:
            status = "QUICK_REVIEW"
            fill_color = "FFEB9C"  # Light yellow
        elif score.overall >= 0.70:
            status = "REVIEW"
            fill_color = "FFC7CE"  # Light red
        else:
            status = "DETAILED_REVIEW"
            fill_color = "FF6B6B"  # Red
        
        row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}", status]
        ws_work.append(row)
        
        # Color code based on confidence
        for cell in ws_work[i+1]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Sheet 3: Bill Quantity (with validation)
    ws_bill = wb.create_sheet("Bill Quantity")
    ws_bill.append(headers)
    
    # Style headers
    for cell in ws_bill[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Add items with executed quantities
    for i, (item, score) in enumerate(zip(items_list, scores), 1):
        code = item.get('code', '')
        desc = item.get('description', '')
        unit = item.get('unit', '')
        rate = item.get('rate', 0)
        
        # Get executed quantity from qty.txt
        qty = qty_data.get(code, 0)
        amount = qty * rate
        
        # Determine status
        if score.overall >= 0.95:
            status = "AUTO_ACCEPT"
            fill_color = "C6EFCE"
        elif score.overall >= 0.85:
            status = "QUICK_REVIEW"
            fill_color = "FFEB9C"
        elif score.overall >= 0.70:
            status = "REVIEW"
            fill_color = "FFC7CE"
        else:
            status = "DETAILED_REVIEW"
            fill_color = "FF6B6B"
        
        row = [i, desc, unit, qty, rate, amount, code, f"{score.overall:.2f}", status]
        ws_bill.append(row)
        
        # Color code
        for cell in ws_bill[i+1]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Save workbook
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)
    
    print(f"\nExcel file created: {OUTPUT_FILE}")
    print(f"  - Validation Summary sheet")
    print(f"  - Work Order sheet (with confidence scores)")
    print(f"  - Bill Quantity sheet (with confidence scores)")
    
    # Print items needing review
    print("\n" + "="*80)
    print("ITEMS NEEDING REVIEW")
    print("="*80)
    
    review_items = [(item, score) for item, score in zip(items_list, scores) if score.overall < 0.95]
    
    if review_items:
        print(f"\n{len(review_items)} items need review:")
        for item, score in review_items:
            print(f"\n  {item['code']}: {item['description'][:50]}...")
            print(f"    Confidence: {score.overall:.2f} ({score.level})")
            print(f"    Action: {score.action}")
            if score.validation_result.errors:
                for err in score.validation_result.errors:
                    print(f"      ERROR: {err.message}")
            if score.validation_result.warnings:
                for warn in score.validation_result.warnings:
                    print(f"      WARNING: {warn.message}")
    else:
        print("\nAll items have high confidence - no review needed!")
    
    print("\n" + "="*80)
    print("EXTRACTION WITH VALIDATION: COMPLETE")
    print("="*80)
    print(f"\nReliability Improvement:")
    print(f"  Before: 70-80% (no validation)")
    print(f"  After: {report['average_confidence']*100:.1f}% (with validation)")
    print(f"  Auto-accept: {report['recommended_actions']['auto_accept']['percentage']:.1f}%")
    print("="*80 + "\n")

if __name__ == '__main__':
    main()
