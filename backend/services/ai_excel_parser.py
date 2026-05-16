"""
AI-Powered Excel Parser Service
Handles haphazard/varied Excel formats from real-world work orders.

Strategy:
  1. Load all sheets with openpyxl (raw, no header assumption)
  2. Heuristic scan: find the data table by locating header-like rows
  3. Map columns intelligently (fuzzy keyword matching)
  4. Extract items with confidence scores
  5. Optionally call LLM for ambiguous column mapping
"""
import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ── Column keyword maps ───────────────────────────────────────────────────────
# Each key = canonical field name, value = list of substrings to match (lowercase)
COLUMN_KEYWORDS: dict[str, list[str]] = {
    "item_no":      ["item no", "item no.", "s.no", "s. no", "sl no", "serial", "bsr", "sno", "#", "item", "ref. bsr", "code"],
    "description":  ["description", "item of work", "particulars", "work", "supply", "detail", "name of item"],
    "unit":         ["unit", "uom", "units"],
    "quantity":     ["quantity", "qty", "nos", "no.", "number", "measure", "upto date", "upto_date", "since last"],
    "rate":         ["rate", "rate rs", "rate (rs)", "unit rate", "price"],
    "amount":       ["amount", "amt", "total", "value", "rs.", "cost"],
    "remark":       ["remark", "remarks", "note", "ref", "bsr ref", "bsr"],
}

# Columns that are numeric
NUMERIC_COLS = {"quantity", "rate", "amount"}


@dataclass
class ColumnMapping:
    canonical: str          # e.g. "quantity"
    col_index: int          # 0-based column index in the sheet
    header_text: str        # original header text found
    confidence: float       # 0.0 – 1.0


@dataclass
class ParsedRow:
    item_no: str = ""
    description: str = ""
    unit: str = ""
    quantity: float = 0.0
    rate: float = 0.0
    amount: float = 0.0
    remark: str = ""
    is_header_row: bool = False
    is_subtotal_row: bool = False
    confidence: float = 1.0
    ai_note: str = ""        # AI suggestion if something looks off


@dataclass
class SheetParseResult:
    sheet_name: str
    header_row_index: int
    column_mappings: list[ColumnMapping]
    rows: list[ParsedRow]
    unmapped_columns: list[str]
    warnings: list[str] = field(default_factory=list)
    total_amount: float = 0.0


@dataclass
class AIParseResult:
    file_id: str
    file_name: str
    sheets: list[SheetParseResult]
    work_order_sheet: Optional[SheetParseResult]
    title_data: dict[str, str]
    ai_suggestions: list[str]
    confidence_overall: float


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean_str(val: Any) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return str(val).strip()


def _to_float(val: Any) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val) if not np.isnan(val) else 0.0
    s = re.sub(r"[^\d.\-]", "", str(val).strip())
    try:
        return float(s)
    except ValueError:
        return 0.0


def _col_score(header_text: str, keywords: list[str]) -> float:
    """Return 0-1 match score between a header cell and keyword list."""
    h = header_text.lower().strip()
    for kw in keywords:
        if kw in h:
            return 1.0 if h == kw else 0.8
    # partial word match
    for kw in keywords:
        for word in kw.split():
            if word in h:
                return 0.5
    return 0.0


def _find_header_row(df: pd.DataFrame, max_scan: int = 30) -> int:
    """
    Scan first `max_scan` rows to find the row most likely to be a column header.
    Heuristic: row with the most non-empty string cells that match known keywords.
    """
    best_row = 0
    best_score = -1

    all_keywords = [kw for kws in COLUMN_KEYWORDS.values() for kw in kws]

    for i in range(min(max_scan, len(df))):
        row = df.iloc[i]
        score = 0
        non_empty = 0
        for cell in row:
            s = _clean_str(cell).lower()
            if not s:
                continue
            non_empty += 1
            for kw in all_keywords:
                if kw in s:
                    score += 1
                    break
        # Penalise rows with very few non-empty cells
        if non_empty >= 2:
            adjusted = score + (non_empty * 0.1)
            if adjusted > best_score:
                best_score = adjusted
                best_row = i

    return best_row


def _map_columns(header_row: pd.Series) -> list[ColumnMapping]:
    """Map each column to a canonical field using keyword scoring."""
    mappings: list[ColumnMapping] = []
    used_canonicals: set[str] = set()

    # Score every column against every canonical field
    scores: list[tuple[float, str, int, str]] = []  # (score, canonical, col_idx, header_text)
    for col_idx, cell in enumerate(header_row):
        h = _clean_str(cell)
        if not h:
            continue
        for canonical, keywords in COLUMN_KEYWORDS.items():
            s = _col_score(h, keywords)
            if s > 0:
                scores.append((s, canonical, col_idx, h))

    # Greedy assignment: highest score first, each canonical used once
    scores.sort(key=lambda x: -x[0])
    for score, canonical, col_idx, header_text in scores:
        if canonical in used_canonicals:
            continue
        mappings.append(ColumnMapping(
            canonical=canonical,
            col_index=col_idx,
            header_text=header_text,
            confidence=score,
        ))
        used_canonicals.add(canonical)

    return mappings


def _is_subtotal_row(row: pd.Series) -> bool:
    """Detect summary/total rows to skip from item list."""
    for cell in row:
        s = _clean_str(cell).lower()
        if s in ("total", "grand total", "sub total", "subtotal", "sum", "net total"):
            return True
        if "tender premium" in s or "add premium" in s:
            return True
    return False


def _extract_title_data(df: pd.DataFrame, header_row_idx: int) -> dict[str, str]:
    """
    Extract key-value metadata from rows above the data table header.
    Looks for label: value patterns.
    """
    title: dict[str, str] = {}
    for i in range(min(header_row_idx, len(df))):
        row = df.iloc[i]
        cells = [_clean_str(c) for c in row if _clean_str(c)]
        if len(cells) >= 2:
            key = cells[0].rstrip(":- ")
            val = cells[1]
            if key and val and len(key) < 80:
                title[key] = val
        elif len(cells) == 1:
            # Could be a section heading — store as flag
            s = cells[0]
            if len(s) < 60 and not s.startswith("FOR CONTRACTORS"):
                title[f"_heading_{i}"] = s
    return title


def _parse_sheet(sheet_name: str, df: pd.DataFrame) -> SheetParseResult:
    """Full parse of one sheet DataFrame."""
    warnings: list[str] = []

    # 1. Find header row
    header_row_idx = _find_header_row(df)
    header_row = df.iloc[header_row_idx]

    # 2. Map columns
    mappings = _map_columns(header_row)
    mapped_canonicals = {m.canonical for m in mappings}
    mapped_col_indices = {m.col_index for m in mappings}

    # Warn about unmapped columns
    unmapped = []
    for col_idx, cell in enumerate(header_row):
        h = _clean_str(cell)
        if h and col_idx not in mapped_col_indices:
            unmapped.append(h)

    if "description" not in mapped_canonicals:
        warnings.append("Could not detect a Description column — using longest text column as fallback")
        # Fallback: pick column with longest average text
        best_col = 0
        best_len = 0
        for col_idx in range(len(df.columns)):
            avg = df.iloc[header_row_idx + 1:, col_idx].apply(
                lambda x: len(_clean_str(x))
            ).mean()
            if avg > best_len:
                best_len = avg
                best_col = col_idx
        mappings.append(ColumnMapping("description", best_col, "?", 0.3))
        mapped_canonicals.add("description")

    if "quantity" not in mapped_canonicals:
        warnings.append("No Quantity column detected — Bill Quantity will be blank (fill manually)")

    # Build lookup: canonical → col_index
    col_map = {m.canonical: m.col_index for m in mappings}

    # 3. Extract title data from rows above header
    title_data = _extract_title_data(df, header_row_idx)

    # 4. Parse data rows
    rows: list[ParsedRow] = []
    total_amount = 0.0

    for i in range(header_row_idx + 1, len(df)):
        row = df.iloc[i]

        # Skip completely empty rows
        non_empty = [_clean_str(c) for c in row if _clean_str(c)]
        if not non_empty:
            continue

        if _is_subtotal_row(row):
            continue

        desc = _clean_str(row.iloc[col_map["description"]]) if "description" in col_map else ""
        item_no = _clean_str(row.iloc[col_map["item_no"]]) if "item_no" in col_map else ""
        unit = _clean_str(row.iloc[col_map["unit"]]) if "unit" in col_map else ""
        qty = _to_float(row.iloc[col_map["quantity"]]) if "quantity" in col_map else 0.0
        rate = _to_float(row.iloc[col_map["rate"]]) if "rate" in col_map else 0.0
        amount = _to_float(row.iloc[col_map["amount"]]) if "amount" in col_map else 0.0
        remark = _clean_str(row.iloc[col_map["remark"]]) if "remark" in col_map else ""

        # PWD-specific: col0 often holds parent item number (1, 2, 3…)
        # col6 (BSR) holds the full dot-notation code (1.5, 1.5.1…)
        # Prefer BSR/item_no col; fall back to col0 for parent rows if it looks like a code
        if not item_no and len(row) > 0:
            col0_val = _clean_str(row.iloc[0])
            # If col0 has a code-like structure (digits and dots)
            if col0_val and (col0_val.replace('.', '').isdigit() or re.match(r'^\d+[\.a-z0-9]*$', col0_val.lower())):
                item_no = col0_val
        # Skip rows with no description AND no item_no
        if not desc and not item_no:
            continue

        # Auto-calculate amount if missing but qty+rate present
        ai_note = ""
        if amount == 0.0 and qty > 0 and rate > 0:
            amount = round(qty * rate)
            ai_note = "Amount auto-calculated from Qty × Rate"

        # Flag suspicious values
        if qty > 0 and rate > 0:
            expected = qty * rate
            if amount > 0 and abs(amount - expected) / max(expected, 1) > 0.05:
                ai_note = f"Amount mismatch: {amount:.0f} vs expected {expected:.0f}"

        parsed = ParsedRow(
            item_no=item_no,
            description=desc,
            unit=unit,
            quantity=qty,
            rate=rate,
            amount=amount,
            remark=remark,
            confidence=0.9 if (desc and qty > 0 and rate > 0) else 0.6,
            ai_note=ai_note,
        )
        rows.append(parsed)
        total_amount += amount

    return SheetParseResult(
        sheet_name=sheet_name,
        header_row_index=header_row_idx,
        column_mappings=mappings,
        rows=rows,
        unmapped_columns=unmapped,
        warnings=warnings,
        total_amount=total_amount,
    )


# ── Public API ────────────────────────────────────────────────────────────────

def parse_excel_ai(
    file_path: Path,
    file_id: str,
    file_name: str,
    use_llm: bool = False,
) -> AIParseResult:
    """
    Main entry point. Parses an Excel file in any format.
    Returns structured AIParseResult with per-sheet data.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    sheet_results: list[SheetParseResult] = []
    ai_suggestions: list[str] = []
    title_data: dict[str, str] = {}

    for sname in sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sname, header=None, engine="openpyxl")
            df = df.replace({np.nan: None})
            result = _parse_sheet(sname, df)
            sheet_results.append(result)

            # Collect title data from first sheet (usually Work Order)
            if not title_data and result.header_row_index > 0:
                title_data = _extract_title_data(df, result.header_row_index)

            for w in result.warnings:
                ai_suggestions.append(f"[{sname}] {w}")

        except Exception as e:
            logger.warning(f"Could not parse sheet '{sname}': {e}")
            ai_suggestions.append(f"Sheet '{sname}' could not be parsed: {e}")

    # Identify the Work Order sheet (most items, or named "Work Order")
    wo_sheet = None
    for sr in sheet_results:
        name_upper = sr.sheet_name.upper()
        if "WORK ORDER" in name_upper or "WO" == name_upper:
            wo_sheet = sr
            break
    if wo_sheet is None and sheet_results:
        wo_sheet = max(sheet_results, key=lambda s: len(s.rows))

    # Overall confidence
    if wo_sheet and wo_sheet.rows:
        avg_conf = sum(r.confidence for r in wo_sheet.rows) / len(wo_sheet.rows)
    else:
        avg_conf = 0.5

    # LLM enhancement (optional, only if API key present)
    if use_llm:
        try:
            _enhance_with_llm(wo_sheet, ai_suggestions)
        except Exception as e:
            logger.warning(f"LLM enhancement skipped: {e}")
            ai_suggestions.append("AI column mapping unavailable — using heuristic mapping")

    return AIParseResult(
        file_id=file_id,
        file_name=file_name,
        sheets=sheet_results,
        work_order_sheet=wo_sheet,
        title_data=title_data,
        ai_suggestions=ai_suggestions,
        confidence_overall=round(avg_conf, 2),
    )


def _enhance_with_llm(sheet: Optional[SheetParseResult], suggestions: list[str]) -> None:
    """
    Optional: call an LLM to improve column mapping for ambiguous cases.
    Uses GROQ_API_KEY or OPENAI_API_KEY from environment.
    """
    import os
    import json

    api_key = os.getenv("GROQ_API_KEY") or os.getenv("OPENAI_API_KEY")
    if not api_key or sheet is None:
        return

    # Build a compact summary of the column headers for the LLM
    headers_summary = ", ".join(
        f'col{m.col_index}="{m.header_text}"' for m in sheet.column_mappings
    )
    unmapped_summary = ", ".join(f'"{h}"' for h in sheet.unmapped_columns)

    prompt = f"""You are analyzing a PWD contractor bill Excel sheet.
Mapped columns: {headers_summary}
Unmapped columns: {unmapped_summary}

For each unmapped column, suggest which canonical field it maps to:
item_no, description, unit, quantity, rate, amount, remark, or IGNORE.
Reply as JSON: {{"col_name": "canonical_field", ...}}"""

    try:
        if os.getenv("GROQ_API_KEY"):
            import httpx
            resp = httpx.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "model": "llama3-8b-8192",
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 200,
                    "temperature": 0,
                },
                timeout=10,
            )
            content = resp.json()["choices"][0]["message"]["content"]
            mapping = json.loads(content)
            for col_name, canonical in mapping.items():
                if canonical != "IGNORE":
                    suggestions.append(f"AI suggests: '{col_name}' → {canonical}")
    except Exception as e:
        logger.debug(f"LLM call failed: {e}")


def ai_parse_result_to_dict(result: AIParseResult) -> dict:
    """Serialize AIParseResult to a JSON-safe dict for the API response."""
    def mapping_to_dict(m: ColumnMapping) -> dict:
        return {
            "canonical": m.canonical,
            "colIndex": m.col_index,
            "headerText": m.header_text,
            "confidence": m.confidence,
        }

    def row_to_dict(r: ParsedRow) -> dict:
        return {
            "itemNo": r.item_no,
            "description": r.description,
            "unit": r.unit,
            "quantity": r.quantity,
            "rate": r.rate,
            "amount": r.amount,
            "remark": r.remark,
            "confidence": r.confidence,
            "aiNote": r.ai_note,
        }

    def sheet_to_dict(s: SheetParseResult) -> dict:
        return {
            "sheetName": s.sheet_name,
            "headerRowIndex": s.header_row_index,
            "columnMappings": [mapping_to_dict(m) for m in s.column_mappings],
            "rows": [row_to_dict(r) for r in s.rows],
            "unmappedColumns": s.unmapped_columns,
            "warnings": s.warnings,
            "totalAmount": s.total_amount,
        }

    wo = result.work_order_sheet
    return {
        "fileId": result.file_id,
        "fileName": result.file_name,
        "sheets": [sheet_to_dict(s) for s in result.sheets],
        "workOrderSheet": sheet_to_dict(wo) if wo else None,
        "titleData": result.title_data,
        "aiSuggestions": result.ai_suggestions,
        "confidenceOverall": result.confidence_overall,
    }
