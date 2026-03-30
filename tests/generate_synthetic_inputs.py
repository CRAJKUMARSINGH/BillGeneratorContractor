#!/usr/bin/env python3
"""
Generate 24 synthetic Excel input files for batch testing.
- 12 WITHOUT extra items (various work types, amounts, bill numbers)
- 12 WITH extra items
All files follow the PWD 4-sheet format: Title | Work Order | Bill Quantity | Extra Items
Output: tests/SYNTHETIC_INPUTS/
"""
import sys, os
sys.path.insert(0, '.')
os.environ.setdefault('ALLOW_INSECURE_SECRET', '1')

from pathlib import Path
from datetime import datetime, timedelta
import random
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path('tests/SYNTHETIC_INPUTS')
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Work item templates ───────────────────────────────────────────────────────
WORK_ITEMS = [
    ("Earthwork excavation in ordinary soil", "cum", 45.0),
    ("Providing and laying PCC 1:4:8", "cum", 4200.0),
    ("Providing and laying RCC M20", "cum", 6800.0),
    ("Brick masonry in CM 1:6", "cum", 3200.0),
    ("Plastering 12mm thick CM 1:6", "sqm", 185.0),
    ("Providing and fixing MS door frame", "each", 2800.0),
    ("Providing and fixing flush door shutter", "sqm", 1450.0),
    ("Painting with OBD two coats", "sqm", 95.0),
    ("Providing and laying DPC 40mm thick", "sqm", 320.0),
    ("Providing and fixing ceramic tiles 300x300", "sqm", 680.0),
    ("Providing and fixing UPVC window", "sqm", 1850.0),
    ("Providing and laying vitrified tiles 600x600", "sqm", 920.0),
    ("Providing and fixing MS grill", "sqm", 1200.0),
    ("Providing and laying IPS flooring", "sqm", 280.0),
    ("Providing and fixing aluminium partition", "sqm", 2400.0),
]

EXTRA_ITEMS = [
    ("E-01", "10.1.1", "Providing and fixing LED light 18W", "Each", 2, 850.0),
    ("E-02", "10.2.3", "Providing and fixing 6A 3-pin socket", "Each", 4, 320.0),
    ("E-03", "10.3.1", "Providing and fixing MCB 32A", "Each", 2, 480.0),
    ("E-04", "11.1.2", "Providing and fixing CPVC pipe 25mm", "Rmt", 15, 185.0),
    ("E-05", "11.2.1", "Providing and fixing ball valve 25mm", "Each", 3, 420.0),
    ("E-06", "12.1.1", "Providing and fixing exhaust fan 300mm", "Each", 2, 1200.0),
]

WORK_NAMES = [
    "Construction of Community Hall at Village Rampur",
    "Repair and Renovation of Government School Building",
    "Construction of Boundary Wall at District Hospital",
    "Providing and Fixing of Water Supply Pipeline",
    "Construction of Drainage System in Ward No. 5",
    "Renovation of PHC Building at Tehsil Headquarters",
    "Construction of Anganwadi Centre at Village Khera",
    "Repair of Road from Main Market to Bus Stand",
    "Construction of Toilet Block at Government College",
    "Providing Street Lighting in Residential Colony",
    "Construction of Check Dam at Nala Crossing",
    "Renovation of Police Station Building",
]

CONTRACTORS = [
    "M/s. Shree Ram Construction Co., Jaipur",
    "M/s. Bajrang Builders, Udaipur",
    "M/s. Shivshakti Traders, Jodhpur",
    "M/s. Krishna Enterprises, Kota",
    "M/s. Rajputana Constructions, Ajmer",
    "M/s. Mewar Infra Pvt. Ltd., Chittorgarh",
]

thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)


def _cell(ws, row, col, value, bold=False, align='left', fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name='Calibri', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = BORDER
    if fill:
        c.fill = fill
    return c


def create_title_sheet(wb, meta: dict):
    ws = wb.create_sheet("Title")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50

    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value = "FOR CONTRACTORS & SUPPLIERS ONLY FOR PAYMENT FOR WORK OR SUPPLIES ACTUALLY MEASURED"
    c.font = Font(bold=True, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    rows = [
        ("Bill Number", meta['bill_no']),
        ("Agreement No.", meta['agreement_no']),
        ("Name of Work", meta['name_of_work']),
        ("Name of Contractor or supplier", meta['contractor']),
        ("Work Order Amount Rs.", meta['wo_amount']),
        ("Date of written order to commence work", meta['date_start']),
        ("St. Date of Completion", meta['date_end']),
        ("Date of actual completion of work", meta.get('date_actual', '')),
        ("Tender Premium %", meta.get('premium_pct', 0)),
        ("Premium Type", meta.get('premium_type', 'Above')),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        _cell(ws, i, 1, k, bold=True, fill=HEADER_FILL)
        _cell(ws, i, 2, v)


def create_work_order_sheet(wb, items: list, meta: dict):
    ws = wb.create_sheet("Work Order")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

    # Header rows (metadata)
    meta_rows = [
        ("Agreement No.", meta['agreement_no']),
        ("Name of Work", meta['name_of_work']),
        ("Name of Contractor or supplier", meta['contractor']),
        ("Work Order Amount Rs.", meta['wo_amount']),
    ]
    for i, (k, v) in enumerate(meta_rows, start=1):
        _cell(ws, i, 1, k, bold=True)
        ws.merge_cells(f'B{i}:G{i}')
        _cell(ws, i, 2, v)

    # Column headers
    headers = ["Item", "Description", "Unit", "Qty as per WO", "Rate", "Amount Rs.", "Remarks"]
    for j, h in enumerate(headers, start=1):
        _cell(ws, 6, j, h, bold=True, align='center', fill=HEADER_FILL)

    # Data rows
    for i, (desc, unit, rate) in enumerate(items, start=7):
        qty = round(random.uniform(10, 200), 2)
        _cell(ws, i, 1, str(i - 6))
        _cell(ws, i, 2, desc)
        _cell(ws, i, 3, unit)
        _cell(ws, i, 4, qty)
        _cell(ws, i, 5, rate)
        _cell(ws, i, 6, round(qty * rate, 2))
        _cell(ws, i, 7, "")


def create_bill_quantity_sheet(wb, items: list, meta: dict):
    ws = wb.create_sheet("Bill Quantity")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    meta_rows = [
        ("Agreement No.", meta['agreement_no']),
        ("Name of Work", meta['name_of_work']),
    ]
    for i, (k, v) in enumerate(meta_rows, start=1):
        _cell(ws, i, 1, k, bold=True)
        ws.merge_cells(f'B{i}:G{i}')
        _cell(ws, i, 2, v)

    headers = ["Item", "Description", "Unit",
               "Qty Since Last Bill", "Qty Upto Date", "Rate", "Remarks"]
    for j, h in enumerate(headers, start=1):
        _cell(ws, 4, j, h, bold=True, align='center', fill=HEADER_FILL)

    for i, (desc, unit, rate) in enumerate(items, start=5):
        qty_upto = round(random.uniform(10, 200), 2)
        qty_since = round(qty_upto * random.uniform(0.3, 1.0), 2)
        _cell(ws, i, 1, str(i - 4))
        _cell(ws, i, 2, desc)
        _cell(ws, i, 3, unit)
        _cell(ws, i, 4, qty_since)
        _cell(ws, i, 5, qty_upto)
        _cell(ws, i, 6, rate)
        _cell(ws, i, 7, "")


def create_extra_items_sheet(wb, extra_items: list, meta: dict):
    ws = wb.create_sheet("Extra Items")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    ws.merge_cells('A1:H1')
    ws['A1'].value = "EXTRA ITEM SLIP"
    ws['A1'].font = Font(bold=True, name='Calibri', size=12)
    ws['A1'].alignment = Alignment(horizontal='center')

    _cell(ws, 2, 1, "Name of Work :-")
    ws.merge_cells('B2:H2')
    _cell(ws, 2, 2, meta['name_of_work'])

    _cell(ws, 3, 1, "Name of Contractor or supplier :")
    ws.merge_cells('B3:H3')
    _cell(ws, 3, 2, meta['contractor'])

    _cell(ws, 4, 1, "Reference to work order or Agreement :")
    ws.merge_cells('B4:H4')
    _cell(ws, 4, 2, meta['agreement_no'])

    headers = ["S.No.", "Ref. BSR No.", "Particulars", "Qty.", "unit", "Rate", "Amount", "Remarks"]
    for j, h in enumerate(headers, start=1):
        _cell(ws, 6, j, h, bold=True, align='center', fill=HEADER_FILL)

    total = 0.0
    for i, (sno, bsr, desc, unit, qty, rate) in enumerate(extra_items, start=7):
        amount = round(qty * rate, 2)
        total += amount
        _cell(ws, i, 1, sno)
        _cell(ws, i, 2, bsr)
        _cell(ws, i, 3, desc)
        _cell(ws, i, 4, qty)
        _cell(ws, i, 5, unit)
        _cell(ws, i, 6, rate)
        _cell(ws, i, 7, amount)
        _cell(ws, i, 8, "Approved")

    last_row = 7 + len(extra_items)
    _cell(ws, last_row, 2, "Total", bold=True)
    _cell(ws, last_row, 7, total, bold=True)


def generate_file(index: int, with_extra: bool, seed: int) -> Path:
    random.seed(seed)

    # Pick random items (5-10 items)
    n_items = random.randint(5, 10)
    items = random.sample(WORK_ITEMS, min(n_items, len(WORK_ITEMS)))

    # Calculate WO amount
    wo_amount = sum(round(random.uniform(10, 200), 2) * rate for _, _, rate in items)
    wo_amount = round(wo_amount * random.uniform(1.05, 1.20))  # slight buffer

    start_date = datetime(2025, random.randint(1, 6), random.randint(1, 28))
    end_date = start_date + timedelta(days=random.randint(90, 365))

    meta = {
        'bill_no': f"BILL-{index:03d}/2025-26",
        'agreement_no': f"{random.randint(10, 99)}/{random.randint(2024, 2025)}-{str(random.randint(25, 26))}",
        'name_of_work': random.choice(WORK_NAMES),
        'contractor': random.choice(CONTRACTORS),
        'wo_amount': wo_amount,
        'date_start': start_date.strftime('%d-%m-%Y'),
        'date_end': end_date.strftime('%d-%m-%Y'),
        'date_actual': '',
        'premium_pct': round(random.uniform(0, 15), 2),
        'premium_type': random.choice(['Above', 'Below']),
    }

    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    create_title_sheet(wb, meta)
    create_work_order_sheet(wb, items, meta)
    create_bill_quantity_sheet(wb, items, meta)

    if with_extra:
        n_extra = random.randint(2, len(EXTRA_ITEMS))
        extra = random.sample(EXTRA_ITEMS, n_extra)
        create_extra_items_sheet(wb, extra, meta)
    else:
        # Still create the sheet but empty (as per PWD format)
        ws = wb.create_sheet("Extra Items")
        ws['A1'] = "EXTRA ITEM SLIP"
        ws['A2'] = "No extra items for this bill."

    suffix = "WithExtra" if with_extra else "NoExtra"
    filename = f"SYNTH_{index:02d}_{suffix}.xlsx"
    path = OUT_DIR / filename
    wb.save(str(path))
    return path


def main():
    print(f"Generating 24 synthetic Excel input files in {OUT_DIR}")
    generated = []

    # 12 WITHOUT extra items (indices 1-12)
    for i in range(1, 13):
        path = generate_file(i, with_extra=False, seed=i * 100)
        generated.append(path)
        print(f"  [{i:2d}/24] {path.name}")

    # 12 WITH extra items (indices 13-24)
    for i in range(13, 25):
        path = generate_file(i, with_extra=True, seed=i * 100)
        generated.append(path)
        print(f"  [{i:2d}/24] {path.name}")

    print(f"\nGenerated {len(generated)} files in {OUT_DIR}")
    return generated


if __name__ == "__main__":
    main()
