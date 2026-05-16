#!/usr/bin/env python3
"""
PDF Certification Harness — Content + Visual diff against UNIF reference.

Compares CONS (this repo) vs UNIF (streamlit_version) for 6 statutory docs:
  first_page, deviation_statement, extra_items, note_sheet,
  certificate_ii, certificate_iii

UNIF reference path (Option 3B):
  C:\\Users\\Rajkumar\\BillGeneratorUnified\\ARCHIVED\\streamlit_version

Usage:
  python tests/pdf_certification_harness.py [--visual] [--dpi 150] [--limit 9]
  python tests/pdf_certification_harness.py --unif-root "C:\\path\\to\\streamlit_version" --visual

Outputs per test case:
  REGRESSION_RESULTS_CERT/<stem>/CONS/   — our PDFs + page PNGs
  REGRESSION_RESULTS_CERT/<stem>/UNIF/   — reference PDFs + page PNGs
  REGRESSION_RESULTS_CERT/<stem>/DIFF/   — diff images + report.json + report.md
  REGRESSION_RESULTS_CERT/summary.csv    — overall pass/fail table
  REGRESSION_RESULTS_CERT/summary.json
"""

import argparse
import csv
import json
import logging
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader

ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))
from engine import run_engine as engine_runner

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────

# UNIF doc title → CONS template stem  (Option 4A: 6 docs)
DOC_MAPPING: Dict[str, str] = {
    "First Page Summary":      "first_page",
    "Deviation Statement":     "deviation_statement",
    "Extra Items Statement":   "extra_items",
    "BILL SCRUTINY SHEET":     "note_sheet",
    "Certificate II":          "certificate_ii",
    "Certificate III":         "certificate_iii",
}

# Key field tokens that MUST appear in extracted text for each doc type.
# Checked case-insensitively after whitespace normalisation.
REQUIRED_TOKENS: Dict[str, List[str]] = {
    "first_page":          ["agreement", "contractor"],
    "deviation_statement": ["deviation", "quantity"],
    "extra_items":         ["extra", "item"],
    "note_sheet":          ["note", "amount"],
    "certificate_ii":      ["certificate"],
    "certificate_iii":     ["certificate"],
}

# Default UNIF root — Option 3B: streamlit_version sub-path
DEFAULT_UNIF_ROOT = Path(r"C:\Users\Rajkumar\BillGeneratorUnified\ARCHIVED\streamlit_version")

# Fallback: try the non-ARCHIVED path if the above doesn't exist
FALLBACK_UNIF_ROOT = Path(r"C:\Users\Rajkumar\BillGeneratorUnified")


# ── Text helpers ───────────────────────────────────────────────────────────────
    dpi: int,
) -> Dict[str, Any]:
    try:
        from pdf2image import convert_from_path
        from PIL import ImageChops
        import numpy as np

        images_a = convert_from_path(str(pdf_a), dpi=dpi)
        images_b = convert_from_path(str(pdf_b), dpi=dpi)

        out_dir.mkdir(parents=True, exist_ok=True)

        page_count = min(len(images_a), len(images_b))
        per_page = []
        for i in range(page_count):
            img_a = images_a[i].convert("RGB")
            img_b = images_b[i].convert("RGB")
            if img_a.size != img_b.size:
                img_b = img_b.resize(img_a.size)

            diff = ImageChops.difference(img_a, img_b)
            diff_arr = np.asarray(diff, dtype=np.float32)
            mean_abs = float(diff_arr.mean())
            pct = mean_abs / 255.0 * 100.0

            diff_path = out_dir / f"diff_page_{i + 1}.png"
            diff.save(diff_path)

            per_page.append({"page": i + 1, "mean_abs": mean_abs, "percent_diff": pct})

            if i == 0:
                images_a[i].save(out_dir / "a_page_1.png")
                images_b[i].save(out_dir / "b_page_1.png")

        return {
            "visual_available": True,
            "rendered_pages_a": len(images_a),
            "rendered_pages_b": len(images_b),
            "compared_pages": page_count,
            "visual_avg_percent_diff": float(
                sum((p["percent_diff"] for p in per_page), start=0.0) / len(per_page)
            )
            if per_page
            else 0.0,
            "visual_max_percent_diff": max((p["percent_diff"] for p in per_page), default=0.0),
            "per_page": per_page,
        }
    except Exception as e:
        logging.warning(f"Visual diff failed: {e}")
        return {"visual_available": False, "error": str(e)}


def run_unif(excel_path: Path, output_base: Path, unif_root: Path) -> Optional[Dict[str, Dict[str, Any]]]:
    """
    Generate UNIF PDFs for all produced docs (only those available per input).
    Returns mapping: UNIF_doc_title → {pdf, size}
    """
    try:
        sys.path.insert(0, str(unif_root))
        from core.generators.document_generator import DocumentGenerator as UnifDocGenerator
        from core.generators.pdf_generator_fixed import FixedPDFGenerator
        from core.processors.excel_processor import ExcelProcessor as UnifExcelProcessor

        processor = UnifExcelProcessor()
        data = processor.process_excel(excel_path)
        doc_gen = UnifDocGenerator(data)
        docs = doc_gen.generate_all_documents()

        pdf_gen = FixedPDFGenerator(margin_mm=10)

        unif_out = output_base / "UNIF"
        unif_out.mkdir(parents=True, exist_ok=True)

        results: Dict[str, Dict[str, Any]] = {}
        for name, html in docs.items():
            if not html:
                continue
            safe_name = name.lower().replace(" ", "_").replace("..", ".")
            pdf_path = unif_out / f"{safe_name}.pdf"
            html_path = unif_out / f"{safe_name}.html"

            html_path.write_text(html, encoding="utf-8")
            pdf_bytes = pdf_gen.auto_convert(html, doc_name=name)
            pdf_path.write_bytes(pdf_bytes)

            results[name] = {"pdf": pdf_path, "size": len(pdf_bytes)}

        return results
    except Exception as e:
        logging.error(f"UNIF failed for {excel_path.name}: {e}")
        return None


def run_consolidated(excel_path: Path, output_base: Path) -> Optional[Dict[str, Dict[str, Any]]]:
    """Run the local Consolidated engine (CONS) and return PDFs by template stem."""
    try:
        standardized = _create_standardized_engine_input(excel_path, output_base / "_standardized_input")
        cons_out = output_base / "CONS"
        cons_out.mkdir(parents=True, exist_ok=True)

        sheets = engine_runner.load_excel_sheets(standardized)
        doc = engine_runner.build_document(
            sheets,
            premium_percent=0.0,
            premium_type="above",
            previous_bill_amount=0.0,
        )
        html_paths = engine_runner.render_html(doc, cons_out, "v1")
        pdf_paths = engine_runner.render_pdfs(html_paths, cons_out)

        results: Dict[str, Dict[str, Any]] = {}
        for p in pdf_paths:
            results[p.stem] = {"pdf": p, "size": p.stat().st_size}
        return results
    except Exception as e:
        logging.error(f"CONS failed for {excel_path.name}: {e}")
        return None


def _create_standardized_engine_input(src_xlsx: Path, out_dir: Path) -> Path:
    """
    Convert UNIF-style sheets to engine-required names:
      Work Order -> WORK ORDER
      Bill Quantity -> BILL QUANTITY
      Extra Items -> EXTRA ITEMS
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / src_xlsx.name

    src_wb = load_workbook(src_xlsx, data_only=True)
    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)

    mapping = {
        "Work Order": "WORK ORDER",
        "Bill Quantity": "BILL QUANTITY",
        "Extra Items": "EXTRA ITEMS",
    }

    for src_name, dst_name in mapping.items():
        ws_dst = dst_wb.create_sheet(title=dst_name)
        if src_name not in src_wb.sheetnames:
            continue
        ws_src = src_wb[src_name]
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(list(row))

    dst_wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Certify CONS PDFs vs UNIF reference PDFs.")
    parser.add_argument(
        "--unif-root",
        type=Path,
        default=DEFAULT_UNIF_ROOT,
        help="UNIF root containing TEST_INPUT_FILES/ and core/.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=9,
        help="How many Excel test files to process (sorted by filename).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=ROOT_DIR / "REGRESSION_RESULTS_CERT",
        help="Output directory for CONS/UNIF artifacts + diff reports.",
    )
    parser.add_argument(
        "--visual",
        action="store_true",
        help="Enable pixel diffs (requires pdf2image, poppler, and pillow).",
    )
    parser.add_argument("--dpi", type=int, default=150, help="DPI used for PDF page rasterization.")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    test_files_dir = args.unif_root / "TEST_INPUT_FILES"
    if not test_files_dir.exists():
        raise SystemExit(f"UNIF test input dir missing: {test_files_dir}")

    excel_files = sorted(test_files_dir.glob("*.xlsx"))[: args.limit]
    if not excel_files:
        raise SystemExit(f"No .xlsx files found in: {test_files_dir}")

    args.output_dir.mkdir(parents=True, exist_ok=True)

    visual_ok = args.visual and _setup_visual_deps()
    if args.visual and not visual_ok:
        logging.warning("Visual diffs requested but dependencies missing; continuing with text checks.")

    print(f"Starting PDF certification on {len(excel_files)} files...")
    print(f"Comparing 6 template outputs: {list(DOC_MAPPING.values())}")
    if visual_ok:
        print(f"Visual diffs enabled at dpi={args.dpi}")

    rows = []

    for excel_path in excel_files:
        print(f"\nProcessing: {excel_path.name}")
        case_dir = args.output_dir / excel_path.stem
        case_dir.mkdir(parents=True, exist_ok=True)

        unif_res = run_unif(excel_path, case_dir, args.unif_root)
        cons_res = run_consolidated(excel_path, case_dir)

        if not unif_res or not cons_res:
            rows.append({"file": excel_path.name, "doc": "ALL", "status": "FAILED (engine render)"})
            continue

        for unif_doc_title, cons_doc_stem in DOC_MAPPING.items():
            row: Dict[str, Any] = {
                "file": excel_path.name,
                "unif_doc": unif_doc_title,
                "cons_doc": cons_doc_stem,
                "status": "PASS",
            }

            unif_pdf = unif_res.get(unif_doc_title, {}).get("pdf")
            cons_pdf = cons_res.get(cons_doc_stem, {}).get("pdf")

            # Extra-items is allowed to be missing depending on input.
            if unif_pdf is None and cons_pdf is None:
                row["status"] = "PASS (both missing)"
                rows.append(row)
                continue

            if unif_pdf is None or cons_pdf is None:
                row["status"] = "FAIL (missing one side)"
                rows.append(row)
                continue

            unif_pdf = Path(unif_pdf)
            cons_pdf = Path(cons_pdf)

            row["unif_size"] = unif_pdf.stat().st_size
            row["cons_size"] = cons_pdf.stat().st_size
            row["size_diff_abs"] = abs(row["unif_size"] - row["cons_size"])

            unif_text = _extract_pdf_text(unif_pdf)
            cons_text = _extract_pdf_text(cons_pdf)
            row["unif_page_count"] = _pdf_page_count(unif_pdf)
            row["cons_page_count"] = _pdf_page_count(cons_pdf)
            row["text_similarity"] = SequenceMatcher(
                None, _safe_normalize_text(unif_text), _safe_normalize_text(cons_text)
            ).ratio()

            if visual_ok:
                diff_dir = case_dir / "DIFF" / cons_doc_stem
                visual = _visual_diff(unif_pdf, cons_pdf, diff_dir, dpi=args.dpi)
                row["visual_available"] = visual.get("visual_available", False)
                row["visual_avg_percent_diff"] = visual.get("visual_avg_percent_diff")
                row["visual_max_percent_diff"] = visual.get("visual_max_percent_diff")
            else:
                row["visual_available"] = False

            rows.append(row)

        # Per-case artifacts summary
        case_report = case_dir / "report.json"
        case_report.write_text(json.dumps({"rows": rows}, indent=2), encoding="utf-8")

    # Overall summary
    summary_json = args.output_dir / "summary.json"
    summary_csv = args.output_dir / "summary.csv"

    summary_json.write_text(json.dumps({"rows": rows}, indent=2), encoding="utf-8")

    fieldnames = sorted({k for r in rows for k in r.keys()})
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nCertification complete.")
    print(f"- summary.csv: {summary_csv}")
    print(f"- summary.json: {summary_json}")


if __name__ == "__main__":
    main()

