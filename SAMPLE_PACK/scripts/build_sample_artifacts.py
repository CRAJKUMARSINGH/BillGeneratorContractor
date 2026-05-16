import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
PACK = ROOT / "SAMPLE_PACK"
INPUTS = PACK / "01_inputs"
UNIF_OUT = PACK / "02_reference_unif_outputs"
CONS_OUT = PACK / "03_consolidated_outputs"
TXT_OUT = PACK / "04_ocr_and_quantity_text"

sys.path.insert(0, str(ROOT))
from engine import run_engine as cons_engine


def generate_cons(xlsx: Path) -> None:
    standardized = create_standardized_engine_input(xlsx)
    case_out = CONS_OUT / xlsx.stem
    case_out.mkdir(parents=True, exist_ok=True)
    sheets = cons_engine.load_excel_sheets(standardized)
    doc = cons_engine.build_document(
        sheets,
        premium_percent=0.0,
        premium_type="above",
        previous_bill_amount=0.0,
    )
    html_paths = cons_engine.render_html(doc, case_out, "v1")
    cons_engine.render_pdfs(html_paths, case_out)

    lines = []
    for i, item in enumerate(doc.items, start=1):
        lines.append(
            f"{i}\t{item.get('serial_no','')}\t{item.get('description','')}\t{item.get('quantity_since_last','')}\t{item.get('rate','')}"
        )
    (TXT_OUT / f"{xlsx.stem}_quantities.txt").write_text("\n".join(lines), encoding="utf-8")


def create_standardized_engine_input(src_xlsx: Path) -> Path:
    """
    Convert UNIF-style sheet names into engine-required sheet names:
      Work Order -> WORK ORDER
      Bill Quantity -> BILL QUANTITY
      Extra Items -> EXTRA ITEMS
    """
    std_dir = INPUTS / "standardized_for_engine"
    std_dir.mkdir(parents=True, exist_ok=True)
    out_path = std_dir / src_xlsx.name

    src_wb = load_workbook(src_xlsx, data_only=True)
    dst_wb = Workbook()
    # Remove default sheet
    default_sheet = dst_wb.active
    dst_wb.remove(default_sheet)

    mapping = {
        "Work Order": "WORK ORDER",
        "Bill Quantity": "BILL QUANTITY",
        "Extra Items": "EXTRA ITEMS",
    }

    for src_name, dst_name in mapping.items():
        ws_dst = dst_wb.create_sheet(title=dst_name)
        if src_name not in src_wb.sheetnames:
            continue
        ws_src = src_wb[src_name]
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(list(row))

    dst_wb.save(out_path)
    return out_path


def generate_unif(xlsx: Path, unif_root: Path) -> None:
    sys.path.insert(0, str(unif_root))
    from core.processors.excel_processor import ExcelProcessor as UnifExcelProcessor
    from core.generators.document_generator import DocumentGenerator as UnifDocGenerator
    from core.generators.pdf_generator_fixed import FixedPDFGenerator

    case_out = UNIF_OUT / xlsx.stem
    case_out.mkdir(parents=True, exist_ok=True)

    data = UnifExcelProcessor().process_excel(xlsx)
    docs = UnifDocGenerator(data).generate_all_documents()
    pdf_gen = FixedPDFGenerator(margin_mm=10)
    for name, html in docs.items():
        if not html:
            continue
        safe = name.lower().replace(" ", "_").replace("..", ".")
        (case_out / f"{safe}.html").write_text(html, encoding="utf-8")
        (case_out / f"{safe}.pdf").write_bytes(pdf_gen.auto_convert(html, doc_name=name))


def main() -> None:
    unif_root = Path(r"C:\Users\Rajkumar\BillGeneratorUnified")
    for d in [UNIF_OUT, CONS_OUT, TXT_OUT]:
        d.mkdir(parents=True, exist_ok=True)

    for xlsx in sorted(INPUTS.glob("*.xlsx")):
        print(f"Building artifacts for {xlsx.name}")
        generate_cons(xlsx)
        generate_unif(xlsx, unif_root)

    print("Done.")


if __name__ == "__main__":
    main()

